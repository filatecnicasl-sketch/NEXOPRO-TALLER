from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form, Depends, Request
from fastapi.responses import Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pymongo import ReturnDocument
import re
import os
import json
import hashlib
import logging
import tempfile
import secrets
import bcrypt
import jwt
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Literal
from datetime import datetime, timezone, date, timedelta
import uuid
import base64
import asyncio
import resend
from twilio.rest import Client as TwilioClient
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from facturae_sign import firmar_facturae, leer_datos_certificado

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY')
JWT_SECRET = os.environ.get('JWT_SECRET', 'change-me')
JWT_ALGORITHM = "HS256"

# --- Almacenamiento de objetos (PDFs originales de documentos) ---
import requests as _requests
STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
APP_NAME = "nexopro"
_storage_key = None


def init_storage():
    global _storage_key
    if _storage_key:
        return _storage_key
    resp = _requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_LLM_KEY}, timeout=30)
    resp.raise_for_status()
    _storage_key = resp.json()["storage_key"]
    return _storage_key


def storage_put(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    resp = _requests.put(f"{STORAGE_URL}/objects/{path}",
                         headers={"X-Storage-Key": key, "Content-Type": content_type},
                         data=data, timeout=120)
    if resp.status_code == 403:
        globals()['_storage_key'] = None
        key = init_storage()
        resp = _requests.put(f"{STORAGE_URL}/objects/{path}",
                             headers={"X-Storage-Key": key, "Content-Type": content_type},
                             data=data, timeout=120)
    resp.raise_for_status()
    return resp.json()


def storage_get(path: str):
    key = init_storage()
    resp = _requests.get(f"{STORAGE_URL}/objects/{path}", headers={"X-Storage-Key": key}, timeout=60)
    if resp.status_code == 403:
        globals()['_storage_key'] = None
        key = init_storage()
        resp = _requests.get(f"{STORAGE_URL}/objects/{path}", headers={"X-Storage-Key": key}, timeout=60)
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")


async def _guardar_pdf(contenido: bytes, filename: str) -> dict:
    """Sube un PDF a object storage y devuelve {pdf_path, pdf_filename}."""
    path = f"{APP_NAME}/pdfs/{new_id()}.pdf"
    res = storage_put(path, contenido, "application/pdf")
    return {"pdf_path": res["path"], "pdf_filename": filename or "documento.pdf"}

# Coste aproximado Gemini 2.5 Flash (USD por token) para estimar el gasto de IA
PRECIO_INPUT_TOKEN = 0.30 / 1_000_000
PRECIO_OUTPUT_TOKEN = 2.50 / 1_000_000
USD_EUR = 0.92

app = FastAPI(title="ERP Base - Clientes, Proveedores y Facturación")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def now_iso():
    return datetime.now(timezone.utc).isoformat()


def new_id():
    return str(uuid.uuid4())


def clean(doc: dict) -> dict:
    doc.pop('_id', None)
    return doc


async def _next_seq(name: str) -> int:
    """Contador atómico en Mongo (seguro ante concurrencia)."""
    doc = await db.counters.find_one_and_update(
        {"_id": name}, {"$inc": {"seq": 1}}, upsert=True, return_document=ReturnDocument.AFTER
    )
    return doc["seq"]


# ---------------------------------------------------------------------------
# AUTENTICACIÓN (Admin JWT) + LICENCIAS
# ---------------------------------------------------------------------------
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email,
               "exp": datetime.now(timezone.utc) + timedelta(hours=12), "type": "access"}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


async def get_current_admin(request: Request) -> dict:
    auth_header = request.headers.get("Authorization", "")
    token = auth_header[7:] if auth_header.startswith("Bearer ") else None
    if not token:
        raise HTTPException(401, "No autenticado")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user or user.get("role") != "admin":
            raise HTTPException(401, "No autorizado")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Sesión expirada")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Token inválido")


class LoginInput(BaseModel):
    email: str
    password: str


@api_router.post("/auth/login")
async def auth_login(data: LoginInput):
    user = await db.users.find_one({"email": data.email.lower().strip()})
    if not user or not verify_password(data.password, user["password_hash"]):
        raise HTTPException(401, "Credenciales incorrectas")
    token = create_access_token(user["id"], user["email"])
    return {"token": token, "user": {"email": user["email"], "name": user.get("name", ""), "role": user["role"]}}


@api_router.get("/auth/me")
async def auth_me(admin: dict = Depends(get_current_admin)):
    return admin


# ---- Licencias ----
class LicenciaInput(BaseModel):
    empresa: str
    email: str = ""
    telefono: str = ""
    precio_mensual: float = 29
    notas: str = ""


def _gen_license_key() -> str:
    return "NEXO-" + secrets.token_hex(4).upper() + "-" + secrets.token_hex(2).upper()


@api_router.post("/admin/licencias")
async def crear_licencia(data: LicenciaInput, admin: dict = Depends(get_current_admin)):
    doc = {
        "id": new_id(),
        "license_key": _gen_license_key(),
        "empresa": data.empresa,
        "email": data.email,
        "telefono": data.telefono,
        "precio_mensual": data.precio_mensual,
        "estado": "activa",
        "ultimo_pago": None,
        "proximo_pago": None,
        "notas": data.notas,
        "created_at": now_iso(),
    }
    await db.licencias.insert_one(dict(doc))
    return clean(doc)


@api_router.get("/admin/licencias")
async def listar_licencias(admin: dict = Depends(get_current_admin)):
    return await db.licencias.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)


@api_router.put("/admin/licencias/{lic_id}")
async def actualizar_licencia(lic_id: str, data: LicenciaInput, admin: dict = Depends(get_current_admin)):
    res = await db.licencias.update_one({"id": lic_id}, {"$set": data.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(404, "Licencia no encontrada")
    return await db.licencias.find_one({"id": lic_id}, {"_id": 0})


@api_router.patch("/admin/licencias/{lic_id}/estado")
async def cambiar_estado_licencia(lic_id: str, estado: str = Form(...), admin: dict = Depends(get_current_admin)):
    if estado not in ("activa", "suspendida"):
        raise HTTPException(400, "Estado no válido")
    res = await db.licencias.update_one({"id": lic_id}, {"$set": {"estado": estado}})
    if res.matched_count == 0:
        raise HTTPException(404, "Licencia no encontrada")
    return await db.licencias.find_one({"id": lic_id}, {"_id": 0})


@api_router.post("/admin/licencias/{lic_id}/pago")
async def registrar_pago(lic_id: str, admin: dict = Depends(get_current_admin)):
    hoy = date.today()
    prox = (hoy.replace(day=1) + timedelta(days=32)).replace(day=1)
    res = await db.licencias.update_one({"id": lic_id}, {"$set": {
        "ultimo_pago": hoy.isoformat(), "proximo_pago": prox.isoformat(), "estado": "activa",
    }})
    if res.matched_count == 0:
        raise HTTPException(404, "Licencia no encontrada")
    return await db.licencias.find_one({"id": lic_id}, {"_id": 0})


@api_router.delete("/admin/licencias/{lic_id}")
async def eliminar_licencia(lic_id: str, admin: dict = Depends(get_current_admin)):
    await db.licencias.delete_one({"id": lic_id})
    return {"ok": True}


@api_router.get("/licencia/verificar/{license_key}")
async def verificar_licencia(license_key: str):
    """Endpoint público usado por la app cliente para saber si está activa."""
    lic = await db.licencias.find_one({"license_key": license_key}, {"_id": 0})
    if not lic:
        return {"valida": False, "estado": "no_encontrada", "mensaje": "Licencia no válida."}
    activa = lic["estado"] == "activa"
    return {
        "valida": activa,
        "estado": lic["estado"],
        "empresa": lic["empresa"],
        "mensaje": "Licencia activa." if activa else "Aplicación desactivada. Contacte con su proveedor.",
    }


# ---------------------------------------------------------------------------
# USUARIOS DEL TALLER (login del ERP) — roles, 2FA, anti fuerza bruta
# ---------------------------------------------------------------------------
import pyotp

APP_ROLES = ("admin", "operario", "recepcion")
MAX_INTENTOS = 5
BLOQUEO_MINUTOS = 15
APP_TOKEN_HORAS = 8


def _politica_password(pwd: str):
    if len(pwd or "") < 8:
        raise HTTPException(400, "La contraseña debe tener al menos 8 caracteres.")
    if not re.search(r"[A-Z]", pwd):
        raise HTTPException(400, "La contraseña debe incluir al menos una mayúscula.")
    if not re.search(r"[0-9]", pwd):
        raise HTTPException(400, "La contraseña debe incluir al menos un número.")


def _app_token(user: dict) -> str:
    payload = {"sub": user["id"], "email": user["email"], "role": user.get("role", "operario"),
               "kind": "app", "exp": datetime.now(timezone.utc) + timedelta(hours=APP_TOKEN_HORAS)}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def _app_user_publico(u: dict) -> dict:
    return {"id": u["id"], "nombre": u.get("nombre", ""), "email": u["email"],
            "role": u.get("role", "operario"), "activo": u.get("activo", True),
            "totp_enabled": bool(u.get("totp_enabled")), "last_login": u.get("last_login")}


async def get_current_app_user(request: Request) -> dict:
    auth_header = request.headers.get("Authorization", "")
    token = auth_header[7:] if auth_header.startswith("Bearer ") else None
    if not token:
        raise HTTPException(401, "No autenticado")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("kind") != "app":
            raise HTTPException(401, "Token inválido")
        user = await db.app_users.find_one({"id": payload["sub"]}, {"_id": 0})
        if not user or not user.get("activo", True):
            raise HTTPException(401, "Usuario no disponible")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Sesión expirada")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Token inválido")


def require_app_role(*roles):
    async def _dep(user: dict = Depends(get_current_app_user)) -> dict:
        if user.get("role") not in roles:
            raise HTTPException(403, "No tienes permisos para esta acción")
        return user
    return _dep


class AppLoginInput(BaseModel):
    email: str
    password: str
    totp_code: Optional[str] = None


@api_router.post("/app/auth/login")
async def app_login(data: AppLoginInput):
    email = data.email.lower().strip()
    user = await db.app_users.find_one({"email": email})
    cred_error = HTTPException(401, "Email o contraseña incorrectos.")
    if not user:
        raise cred_error
    if not user.get("activo", True):
        raise HTTPException(403, "Tu cuenta está desactivada. Contacta con el administrador.")
    # Bloqueo por fuerza bruta
    bloqueo = user.get("locked_until")
    if bloqueo and datetime.fromisoformat(bloqueo) > datetime.now(timezone.utc):
        restan = int((datetime.fromisoformat(bloqueo) - datetime.now(timezone.utc)).total_seconds() // 60) + 1
        raise HTTPException(423, f"Cuenta bloqueada por intentos fallidos. Inténtalo de nuevo en {restan} min.")
    if not verify_password(data.password, user["password_hash"]):
        intentos = int(user.get("failed_attempts", 0)) + 1
        upd = {"failed_attempts": intentos}
        if intentos >= MAX_INTENTOS:
            upd["locked_until"] = (datetime.now(timezone.utc) + timedelta(minutes=BLOQUEO_MINUTOS)).isoformat()
            upd["failed_attempts"] = 0
        await db.app_users.update_one({"id": user["id"]}, {"$set": upd})
        if upd.get("locked_until"):
            raise HTTPException(423, f"Demasiados intentos. Cuenta bloqueada {BLOQUEO_MINUTOS} min.")
        raise cred_error
    # 2FA
    if user.get("totp_enabled"):
        if not data.totp_code:
            return {"requires_2fa": True}
        if not pyotp.TOTP(user["totp_secret"]).verify(data.totp_code.strip(), valid_window=1):
            raise HTTPException(401, "Código de verificación incorrecto.")
    await db.app_users.update_one({"id": user["id"]}, {"$set": {
        "failed_attempts": 0, "locked_until": None, "last_login": now_iso()}})
    return {"token": _app_token(user), "user": _app_user_publico(user),
            "must_change_password": bool(user.get("must_change_password"))}


@api_router.get("/app/auth/me")
async def app_me(user: dict = Depends(get_current_app_user)):
    out = _app_user_publico(user)
    out["must_change_password"] = bool(user.get("must_change_password"))
    return out


class CambioPasswordInput(BaseModel):
    actual: str
    nueva: str


@api_router.post("/app/auth/change-password")
async def app_change_password(data: CambioPasswordInput, user: dict = Depends(get_current_app_user)):
    if not verify_password(data.actual, user["password_hash"]):
        raise HTTPException(400, "La contraseña actual no es correcta.")
    _politica_password(data.nueva)
    await db.app_users.update_one({"id": user["id"]}, {"$set": {
        "password_hash": hash_password(data.nueva), "must_change_password": False}})
    return {"ok": True}


# ---- 2FA (TOTP) ----
@api_router.post("/app/auth/2fa/setup")
async def app_2fa_setup(user: dict = Depends(get_current_app_user)):
    secret = pyotp.random_base32()
    await db.app_users.update_one({"id": user["id"]}, {"$set": {"totp_secret": secret, "totp_enabled": False}})
    uri = pyotp.TOTP(secret).provisioning_uri(name=user["email"], issuer_name="NexoPro")
    return {"secret": secret, "otpauth_uri": uri}


class TotpCodeInput(BaseModel):
    code: str


@api_router.post("/app/auth/2fa/enable")
async def app_2fa_enable(data: TotpCodeInput, user: dict = Depends(get_current_app_user)):
    if not user.get("totp_secret"):
        raise HTTPException(400, "Primero genera el código QR (setup).")
    if not pyotp.TOTP(user["totp_secret"]).verify(data.code.strip(), valid_window=1):
        raise HTTPException(400, "Código incorrecto. Revisa la app de autenticación.")
    await db.app_users.update_one({"id": user["id"]}, {"$set": {"totp_enabled": True}})
    return {"ok": True}


@api_router.post("/app/auth/2fa/disable")
async def app_2fa_disable(data: TotpCodeInput, user: dict = Depends(get_current_app_user)):
    if user.get("totp_enabled") and not pyotp.TOTP(user.get("totp_secret", "")).verify(data.code.strip(), valid_window=1):
        raise HTTPException(400, "Código incorrecto.")
    await db.app_users.update_one({"id": user["id"]}, {"$set": {"totp_enabled": False, "totp_secret": None}})
    return {"ok": True}


# ---- Gestión de usuarios (solo admin) ----
class AppUsuarioInput(BaseModel):
    nombre: str = ""
    email: str
    password: Optional[str] = None
    role: str = "operario"
    activo: bool = True


@api_router.get("/app/usuarios")
async def app_listar_usuarios(_: dict = Depends(require_app_role("admin"))):
    docs = await db.app_users.find({}, {"_id": 0}).sort("created_at", 1).to_list(500)
    return [_app_user_publico(d) for d in docs]


@api_router.post("/app/usuarios")
async def app_crear_usuario(data: AppUsuarioInput, _: dict = Depends(require_app_role("admin"))):
    email = data.email.lower().strip()
    if not email:
        raise HTTPException(400, "El email es obligatorio.")
    if data.role not in APP_ROLES:
        raise HTTPException(400, "Rol no válido.")
    if await db.app_users.find_one({"email": email}):
        raise HTTPException(409, "Ya existe un usuario con ese email.")
    if not data.password:
        raise HTTPException(400, "La contraseña es obligatoria.")
    _politica_password(data.password)
    doc = {"id": new_id(), "nombre": data.nombre, "email": email,
           "password_hash": hash_password(data.password), "role": data.role,
           "activo": data.activo, "failed_attempts": 0, "locked_until": None,
           "totp_secret": None, "totp_enabled": False, "must_change_password": True,
           "last_login": None, "created_at": now_iso()}
    await db.app_users.insert_one(dict(doc))
    return _app_user_publico(doc)


@api_router.put("/app/usuarios/{uid}")
async def app_editar_usuario(uid: str, data: AppUsuarioInput, admin: dict = Depends(require_app_role("admin"))):
    user = await db.app_users.find_one({"id": uid})
    if not user:
        raise HTTPException(404, "Usuario no encontrado")
    if data.role not in APP_ROLES:
        raise HTTPException(400, "Rol no válido.")
    # No permitir que el admin se quite a sí mismo el rol/acceso y quede sin administradores
    if user["id"] == admin["id"] and (data.role != "admin" or not data.activo):
        raise HTTPException(400, "No puedes quitarte a ti mismo el rol de administrador ni desactivarte.")
    upd = {"nombre": data.nombre, "role": data.role, "activo": data.activo}
    await db.app_users.update_one({"id": uid}, {"$set": upd})
    return _app_user_publico({**user, **upd})


class ResetPasswordInput(BaseModel):
    nueva: str


@api_router.post("/app/usuarios/{uid}/reset-password")
async def app_reset_password(uid: str, data: ResetPasswordInput, _: dict = Depends(require_app_role("admin"))):
    user = await db.app_users.find_one({"id": uid})
    if not user:
        raise HTTPException(404, "Usuario no encontrado")
    _politica_password(data.nueva)
    await db.app_users.update_one({"id": uid}, {"$set": {
        "password_hash": hash_password(data.nueva), "must_change_password": True,
        "failed_attempts": 0, "locked_until": None}})
    return {"ok": True}


@api_router.delete("/app/usuarios/{uid}")
async def app_eliminar_usuario(uid: str, admin: dict = Depends(require_app_role("admin"))):
    if uid == admin["id"]:
        raise HTTPException(400, "No puedes eliminar tu propia cuenta.")
    total_admins = await db.app_users.count_documents({"role": "admin", "activo": True})
    target = await db.app_users.find_one({"id": uid})
    if target and target.get("role") == "admin" and total_admins <= 1:
        raise HTTPException(400, "Debe existir al menos un administrador.")
    await db.app_users.delete_one({"id": uid})
    return {"ok": True}




# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
class Contacto(BaseModel):
    id: str = Field(default_factory=new_id)
    tipo: Literal['cliente', 'proveedor']
    nombre: str
    nif: str = ""
    email: str = ""
    telefono: str = ""
    direccion: str = ""
    ciudad: str = ""
    codigo_postal: str = ""
    pais: str = "España"
    iban: str = ""
    banco: str = ""
    swift: str = ""
    direccion_entrega: str = ""
    ciudad_entrega: str = ""
    cp_entrega: str = ""
    es_publica: bool = False
    dir3_oficina_contable: str = ""
    dir3_organo_gestor: str = ""
    dir3_unidad_tramitadora: str = ""
    notas: str = ""
    created_at: str = Field(default_factory=now_iso)


class ContactoInput(BaseModel):
    tipo: Literal['cliente', 'proveedor']
    nombre: str
    nif: str = ""
    email: str = ""
    telefono: str = ""
    direccion: str = ""
    ciudad: str = ""
    codigo_postal: str = ""
    pais: str = "España"
    iban: str = ""
    banco: str = ""
    swift: str = ""
    direccion_entrega: str = ""
    ciudad_entrega: str = ""
    cp_entrega: str = ""
    es_publica: bool = False
    dir3_oficina_contable: str = ""
    dir3_organo_gestor: str = ""
    dir3_unidad_tramitadora: str = ""
    notas: str = ""


class LineaItem(BaseModel):
    codigo_proveedor: str = ""
    descripcion: str = ""
    cantidad: float = 1
    unidad: str = "ud"
    precio_unitario: float = 0
    descuento: float = 0        # porcentaje
    tipo_iva: float = 21        # porcentaje
    base: float = 0
    cuota_iva: float = 0
    total: float = 0


def calcular_lineas(lineas: List[dict]):
    resultado = []
    base_total = 0.0
    iva_total = 0.0
    for l in lineas:
        cantidad = float(l.get('cantidad', 0) or 0)
        precio = float(l.get('precio_unitario', 0) or 0)
        descuento = float(l.get('descuento', 0) or 0)
        tipo_iva = float(l.get('tipo_iva', 0) or 0)
        base = round(cantidad * precio * (1 - descuento / 100), 2)
        cuota = round(base * tipo_iva / 100, 2)
        total = round(base + cuota, 2)
        base_total += base
        iva_total += cuota
        resultado.append({
            'codigo_proveedor': l.get('codigo_proveedor', ''),
            'descripcion': l.get('descripcion', ''),
            'cantidad': cantidad,
            'unidad': l.get('unidad', 'ud'),
            'precio_unitario': precio,
            'descuento': descuento,
            'tipo_iva': tipo_iva,
            'base': base,
            'cuota_iva': cuota,
            'total': total,
        })
    base_total = round(base_total, 2)
    iva_total = round(iva_total, 2)
    return resultado, base_total, iva_total, round(base_total + iva_total, 2)


# ---- Documento (Pedidos / Albaranes) ----
class DocumentoInput(BaseModel):
    tipo_operacion: Literal['venta', 'compra'] = 'venta'
    serie: str = ""
    contacto_id: str = ""
    contacto_nombre: str = ""
    contacto_nif: str = ""
    fecha: str = Field(default_factory=lambda: date.today().isoformat())
    estado: str = "borrador"
    lineas: List[LineaItem] = []
    pdf_path: str = ""
    pdf_filename: str = ""
    vehiculo_id: str = ""
    vehiculo_matricula: str = ""
    notas: str = ""


# ---- Facturas Emitidas ----
class FacturaEmitidaInput(BaseModel):
    serie: str = "A"
    cliente_id: str = ""
    cliente_nombre: str = ""
    cliente_nif: str = ""
    fecha_expedicion: str = Field(default_factory=lambda: date.today().isoformat())
    lineas: List[LineaItem] = []
    estado: str = "emitida"
    forma_pago: str = "Transferencia"
    notas: str = ""


# ---- Facturas Recibidas ----
class FacturaRecibidaInput(BaseModel):
    numero_proveedor: str = ""
    proveedor_id: str = ""
    proveedor_nombre: str = ""
    proveedor_nif: str = ""
    fecha: str = Field(default_factory=lambda: date.today().isoformat())
    lineas: List[LineaItem] = []
    estado: str = "pendiente"
    origen: str = "manual"
    forma_pago: str = "Transferencia"
    pdf_base64: str = ""
    pdf_path: str = ""
    pdf_filename: str = ""
    albaranes_ids: List[str] = []
    vehiculo_id: str = ""
    vehiculo_matricula: str = ""
    notas: str = ""


# ---------------------------------------------------------------------------
# CONTACTOS  (Clientes / Proveedores)
# ---------------------------------------------------------------------------
@api_router.post("/contactos")
async def crear_contacto(data: ContactoInput):
    contacto = Contacto(**data.model_dump())
    await db.contactos.insert_one(contacto.model_dump())
    return contacto.model_dump()


_IMPORT_COLMAP = {
    "nombre": ["nombre", "razon social", "razón social", "razonsocial", "cliente", "empresa", "denominacion"],
    "nif": ["nif", "cif", "dni", "nif/cif", "nifcif", "documento", "nif cif"],
    "email": ["email", "correo", "correo electronico", "correo electrónico", "e-mail", "mail"],
    "telefono": ["telefono", "teléfono", "tel", "movil", "móvil", "telefono1", "movil1"],
    "direccion": ["direccion", "dirección", "domicilio", "direccion fiscal"],
    "ciudad": ["ciudad", "poblacion", "población", "localidad"],
    "codigo_postal": ["cp", "codigo postal", "código postal", "c.p.", "codigopostal", "cod postal"],
    "pais": ["pais", "país"],
    "iban": ["iban", "cuenta", "cuenta bancaria"],
    "notas": ["notas", "observaciones", "obs", "comentarios"],
}


def _norm_header(s):
    import unicodedata
    return "".join(c for c in unicodedata.normalize("NFKD", str(s or "").strip().lower()) if not unicodedata.combining(c))


@api_router.post("/contactos/importar")
async def importar_contactos(tipo: str = Form("cliente"), file: UploadFile = File(...)):
    import io
    from openpyxl import load_workbook
    if tipo not in ("cliente", "proveedor"):
        tipo = "cliente"
    contenido = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "El archivo no es un Excel (.xlsx) válido")
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        wb.close()
        raise HTTPException(400, "El archivo está vacío")
    idx = {}
    for i, h in enumerate(headers or []):
        hn = _norm_header(h)
        for field, syn in _IMPORT_COLMAP.items():
            if hn == field or hn in syn:
                idx[i] = field
                break
    if "nombre" not in idx.values():
        wb.close()
        raise HTTPException(400, "No se encontró la columna 'Nombre'. Descarga la plantilla y respeta las cabeceras.")
    creados, omitidos, errores = 0, 0, []
    fila = 1
    for row in rows:
        fila += 1
        d = {"tipo": tipo, "pais": "España"}
        for i, val in enumerate(row):
            f = idx.get(i)
            if f and val is not None and str(val).strip():
                d[f] = str(val).strip()
        if not d.get("nombre"):
            omitidos += 1
            continue
        try:
            await db.contactos.insert_one(Contacto(**d).model_dump())
            creados += 1
        except Exception as e:
            errores.append(f"Fila {fila}: {str(e)[:80]}")
    wb.close()
    return {"creados": creados, "omitidos": omitidos, "errores": errores[:20], "total_errores": len(errores)}


@api_router.get("/contactos/plantilla-excel")
async def plantilla_contactos_excel():
    import io
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(["Nombre", "NIF/CIF", "Email", "Teléfono", "Dirección", "Ciudad", "Código Postal", "País", "IBAN", "Notas"])
    ws.append(["Cliente Ejemplo SL", "B12345678", "cliente@ejemplo.com", "600123456", "Calle Mayor 1", "Madrid", "28001", "España", "", ""])
    for i, w in enumerate([28, 14, 24, 14, 26, 16, 12, 12, 24, 20], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": 'attachment; filename="plantilla_clientes.xlsx"'})



@api_router.get("/contactos")
async def listar_contactos(tipo: Optional[str] = None):
    q = {}
    if tipo:
        q['tipo'] = tipo
    docs = await db.contactos.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return docs


@api_router.get("/contactos/{contacto_id}")
async def obtener_contacto(contacto_id: str):
    doc = await db.contactos.find_one({"id": contacto_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Contacto no encontrado")
    return doc


@api_router.put("/contactos/{contacto_id}")
async def actualizar_contacto(contacto_id: str, data: ContactoInput):
    res = await db.contactos.update_one({"id": contacto_id}, {"$set": data.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(404, "Contacto no encontrado")
    return await db.contactos.find_one({"id": contacto_id}, {"_id": 0})


@api_router.delete("/contactos/{contacto_id}")
async def eliminar_contacto(contacto_id: str):
    await db.contactos.delete_one({"id": contacto_id})
    return {"ok": True}


# ---------------------------------------------------------------------------
# ARTÍCULOS / PRODUCTOS
# ---------------------------------------------------------------------------
class ArticuloInput(BaseModel):
    nombre: str
    descripcion: str = ""
    precio: float = 0
    tipo_iva: float = 21
    unidad: str = "ud"
    codigo_proveedor: str = ""
    codigo_barras: str = ""
    notas: str = ""


async def _next_articulo_ref() -> str:
    n = await _next_seq("articulo_ref")
    return f"ART-{n:06d}"


@api_router.post("/articulos")
async def crear_articulo(data: ArticuloInput):
    d = data.model_dump()
    d["referencia"] = await _next_articulo_ref()
    doc = {"id": new_id(), **d, "nombre_lower": d["nombre"].strip().lower(),
           "origenes": [], "auto": False, "created_at": now_iso()}
    await db.articulos.insert_one(dict(doc))
    return clean(doc)


@api_router.get("/articulos")
async def listar_articulos():
    return await db.articulos.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)


@api_router.put("/articulos/{articulo_id}")
async def actualizar_articulo(articulo_id: str, data: ArticuloInput):
    existing = await db.articulos.find_one({"id": articulo_id})
    if not existing:
        raise HTTPException(404, "Artículo no encontrado")
    d = data.model_dump()
    d["nombre_lower"] = d["nombre"].strip().lower()
    d["referencia"] = existing.get("referencia", "")  # la referencia es automática, no editable
    await db.articulos.update_one({"id": articulo_id}, {"$set": d})
    return await db.articulos.find_one({"id": articulo_id}, {"_id": 0})


@api_router.delete("/articulos/{articulo_id}")
async def eliminar_articulo(articulo_id: str):
    await db.articulos.delete_one({"id": articulo_id})
    return {"ok": True}


async def registrar_articulos_entrada(lineas: list, origen: dict):
    """Da de alta (o actualiza) los artículos que aparecen en un documento de entrada,
    guardando en cada artículo a qué documento(s) corresponde."""
    for l in lineas:
        nombre = (l.get("descripcion") or "").strip()
        if not nombre:
            continue
        entry = {
            "tipo": origen["tipo"],
            "documento_numero": origen["numero"],
            "documento_id": origen["id"],
            "fecha": origen.get("fecha", ""),
            "proveedor": origen.get("proveedor", ""),
            "precio": l.get("precio_unitario", 0),
            "cantidad": l.get("cantidad", 0),
            "codigo_proveedor": l.get("codigo_proveedor", ""),
        }
        cod_prov = l.get("codigo_proveedor", "")
        existing = await db.articulos.find_one({"nombre_lower": nombre.lower()})
        if existing:
            origenes = existing.get("origenes", [])
            if not any(o.get("documento_id") == origen["id"] for o in origenes):
                origenes.append(entry)
            update = {
                "origenes": origenes,
                "precio": l.get("precio_unitario", existing.get("precio", 0)),
                "tipo_iva": l.get("tipo_iva", existing.get("tipo_iva", 21)),
            }
            if cod_prov and not existing.get("codigo_proveedor"):
                update["codigo_proveedor"] = cod_prov
            await db.articulos.update_one({"id": existing["id"]}, {"$set": update})
        else:
            art = {
                "id": new_id(), "referencia": await _next_articulo_ref(), "nombre": nombre,
                "nombre_lower": nombre.lower(), "descripcion": "",
                "precio": l.get("precio_unitario", 0), "tipo_iva": l.get("tipo_iva", 21),
                "unidad": "ud", "codigo_proveedor": cod_prov,
                "codigo_barras": "", "notas": "", "origenes": [entry], "auto": True,
                "created_at": now_iso(),
            }
            await db.articulos.insert_one(dict(art))


async def ensure_cliente(nombre: str, nif: str = "") -> Optional[dict]:
    """Busca un cliente por NIF o nombre; si no existe, lo da de alta automáticamente."""
    nombre = (nombre or "").strip()
    if not nombre:
        return None
    query = {"tipo": "cliente"}
    existing = None
    if nif:
        existing = await db.contactos.find_one({"tipo": "cliente", "nif": nif})
    if not existing:
        existing = await db.contactos.find_one({
            "tipo": "cliente",
            "nombre": {"$regex": f"^{re.escape(nombre)}$", "$options": "i"},
        })
    if existing:
        return existing
    nuevo = Contacto(tipo="cliente", nombre=nombre, nif=nif, notas="Alta automática desde documento de venta")
    doc = nuevo.model_dump()
    await db.contactos.insert_one(dict(doc))
    return doc


async def ensure_proveedor(nombre: str, nif: str = "") -> Optional[dict]:
    """Busca un proveedor por NIF o nombre; si no existe, lo da de alta automáticamente."""
    nombre = (nombre or "").strip()
    if not nombre:
        return None
    existing = None
    if nif:
        existing = await db.contactos.find_one({"tipo": "proveedor", "nif": nif})
    if not existing:
        existing = await db.contactos.find_one({
            "tipo": "proveedor",
            "nombre": {"$regex": f"^{re.escape(nombre)}$", "$options": "i"},
        })
    if existing:
        return existing
    nuevo = Contacto(tipo="proveedor", nombre=nombre, nif=nif, notas="Alta automática desde documento de compra")
    doc = nuevo.model_dump()
    await db.contactos.insert_one(dict(doc))
    return doc


# ---------------------------------------------------------------------------
# AJUSTES (Series de documentos, contadores y datos de empresa)
# ---------------------------------------------------------------------------
DEFAULT_EMPRESA = {
    "nombre": "", "nif": "", "direccion": "", "codigo_postal": "",
    "ciudad": "", "telefono": "", "email": "", "iban": "", "logo": "",
}

DEFAULT_NOTIF = {
    "email": {"activo": False, "api_key": "", "from_email": "", "from_nombre": ""},
    "whatsapp": {"activo": False, "account_sid": "", "auth_token": "", "from_number": ""},
    "recordatorios": {
        "activo": False,
        "horas_antes": 24,
        "canal": "email",  # email | whatsapp | ambos
        "email_asunto": "Recordatorio de su cita en {empresa}",
        "email_cuerpo": ("Hola {cliente},\n\nLe recordamos su cita en {empresa} el {fecha} a las {hora}.\n"
                         "Vehículo: {matricula}\nMotivo: {motivo}\n\n"
                         "Puede confirmar o cancelar su cita aquí: {enlace}\n\nUn saludo."),
        "whatsapp_texto": ("Hola {cliente}, le recordamos su cita en {empresa} el {fecha} a las {hora}. "
                           "Vehículo: {matricula}. {motivo}\nConfirmar o cancelar: {enlace}"),
    },
}
NOTIF_SECRETOS = [("email", "api_key"), ("whatsapp", "auth_token")]
MASK = "••••••••"


def _merge_notif(cfg: dict) -> dict:
    """Combina la config guardada con los valores por defecto (deep merge simple)."""
    n = cfg.get("notificaciones") or {}
    out = {}
    for seccion, defval in DEFAULT_NOTIF.items():
        out[seccion] = {**defval, **(n.get(seccion) or {})}
    return out


async def _get_ajustes() -> dict:
    cfg = await db.ajustes.find_one({"_id": "config"})
    if not cfg:
        count_a = await db.facturas_emitidas.count_documents({"serie": "A"})
        cfg = {
            "_id": "config",
            "empresa": dict(DEFAULT_EMPRESA),
            "notificaciones": dict(DEFAULT_NOTIF),
            "series_venta": [{"id": new_id(), "nombre": "A", "por_defecto": True,
                              "contadores": {"presupuestos": 1, "facturas": count_a + 1, "pedidos": 1, "albaranes": 1}}],
            "series_compra": [{"id": new_id(), "nombre": "C", "por_defecto": True,
                               "contadores": {"pedidos": 1, "albaranes": 1}}],
            "created_at": now_iso(), "updated_at": now_iso(),
        }
        await db.ajustes.insert_one(dict(cfg))
    return cfg


async def _siguiente_contador(ambito: str, tipo_doc: str, serie: str):
    """Próximo número de la serie; incrementa el contador. None si la serie no existe."""
    await _get_ajustes()
    key = "series_venta" if ambito == "venta" else "series_compra"
    res = await db.ajustes.find_one_and_update(
        {"_id": "config", f"{key}.nombre": serie},
        {"$inc": {f"{key}.$.contadores.{tipo_doc}": 1}},
        return_document=ReturnDocument.BEFORE,
    )
    if not res:
        return None
    for s in res.get(key, []):
        if s.get("nombre") == serie:
            return int((s.get("contadores") or {}).get(tipo_doc, 1) or 1)
    return None


class AjustesInput(BaseModel):
    empresa: dict = {}
    series_venta: List[dict] = []
    series_compra: List[dict] = []
    notificaciones: Optional[dict] = None
    app_url: Optional[str] = None
    modulo_inicio: Optional[str] = None
    formato_hoja_entrada: Optional[str] = None


def _norm_series(series: list, tipos: list) -> list:
    out, vistos = [], set()
    for s in series:
        nombre = (s.get("nombre") or "").strip().upper()
        if not nombre or nombre in vistos:
            continue
        vistos.add(nombre)
        cont = s.get("contadores") or {}
        out.append({
            "id": s.get("id") or new_id(),
            "nombre": nombre,
            "por_defecto": bool(s.get("por_defecto")),
            "contadores": {t: max(1, int(cont.get(t, 1) or 1)) for t in tipos},
        })
    if out and not any(s["por_defecto"] for s in out):
        out[0]["por_defecto"] = True
    return out


@api_router.get("/ajustes")
async def obtener_ajustes():
    cfg = dict(await _get_ajustes())
    cfg.setdefault("modulo_inicio", "panel")
    cfg.setdefault("formato_hoja_entrada", None)
    notif = _merge_notif(cfg)
    # Enmascara secretos y expone si están configurados
    for seccion, campo in NOTIF_SECRETOS:
        val = notif.get(seccion, {}).get(campo) or ""
        notif[seccion][f"{campo}_set"] = bool(val)
        notif[seccion][campo] = MASK if val else ""
    cfg["notificaciones"] = notif
    # Facturae/FACe: exponer solo metadatos, nunca el certificado ni la contraseña
    fe = cfg.get("facturae") or {}
    cfg["facturae"] = {
        "entorno": fe.get("entorno") or "pruebas",
        "proveedor_email": fe.get("proveedor_email") or "",
        "cert_configurado": bool(fe.get("cert_p12_base64")),
        "cert_titular": fe.get("cert_titular") or "",
        "cert_valido_desde": fe.get("cert_valido_desde") or "",
        "cert_valido_hasta": fe.get("cert_valido_hasta") or "",
        "cert_emisor": fe.get("cert_emisor") or "",
    }
    return clean(cfg)


@api_router.put("/ajustes")
async def guardar_ajustes(data: AjustesInput):
    prev = await _get_ajustes()
    empresa = {**DEFAULT_EMPRESA, **(data.empresa or {})}
    sv = _norm_series(data.series_venta, ["presupuestos", "facturas", "pedidos", "albaranes"])
    sc = _norm_series(data.series_compra, ["pedidos", "albaranes"])
    if not sv:
        sv = [{"id": new_id(), "nombre": "A", "por_defecto": True,
               "contadores": {"presupuestos": 1, "facturas": 1, "pedidos": 1, "albaranes": 1}}]
    if not sc:
        sc = [{"id": new_id(), "nombre": "C", "por_defecto": True,
               "contadores": {"pedidos": 1, "albaranes": 1}}]
    set_doc = {"empresa": empresa, "series_venta": sv, "series_compra": sc, "updated_at": now_iso()}
    if data.app_url:
        set_doc["app_url"] = data.app_url.rstrip("/")
    if data.modulo_inicio in ("panel", "taller"):
        set_doc["modulo_inicio"] = data.modulo_inicio
    if data.formato_hoja_entrada is not None:
        set_doc["formato_hoja_entrada"] = data.formato_hoja_entrada or None
    if data.notificaciones is not None:
        prev_notif = _merge_notif(prev)
        notif = {}
        for seccion, defval in DEFAULT_NOTIF.items():
            incoming = (data.notificaciones.get(seccion) or {})
            merged = {**defval, **prev_notif.get(seccion, {}), **incoming}
            # limpia claves auxiliares del GET
            merged = {k: v for k, v in merged.items() if not k.endswith("_set")}
            notif[seccion] = merged
        # Mantiene los secretos previos si llega el valor enmascarado o vacío-no-enviado
        for seccion, campo in NOTIF_SECRETOS:
            entrante = (data.notificaciones.get(seccion) or {}).get(campo)
            if entrante is None or entrante == MASK:
                notif[seccion][campo] = prev_notif.get(seccion, {}).get(campo, "")
        set_doc["notificaciones"] = notif
    await db.ajustes.update_one({"_id": "config"}, {"$set": set_doc}, upsert=True)
    return await obtener_ajustes()


# ---------------------------------------------------------------------------
# FACTURAE FASE 2 — Certificado digital (.p12/.pfx) + firma XAdES-EPES + FACe
# ---------------------------------------------------------------------------
FACE_WSDL = {
    "pruebas": "https://se-face-webservice.redsara.es/facturasspp2?wsdl",
    "produccion": "https://webservice.face.gob.es/facturasspp2?wsdl",
}


@api_router.post("/facturae/certificado")
async def subir_certificado_facturae(
    file: UploadFile = File(...),
    password: str = Form(...),
    entorno: str = Form("pruebas"),
    proveedor_email: str = Form(""),
):
    """Sube y valida el certificado .p12/.pfx para la firma de facturas electrónicas."""
    raw = await file.read()
    if len(raw) > 5 * 1024 * 1024:
        raise HTTPException(400, "El certificado no puede superar 5 MB")
    try:
        datos = leer_datos_certificado(raw, password)
    except Exception:
        raise HTTPException(400, "No se pudo abrir el certificado. Revisa el fichero .p12/.pfx y la contraseña.")
    await _get_ajustes()
    fe = {
        "cert_p12_base64": base64.b64encode(raw).decode(),
        "cert_password": password,
        "entorno": entorno if entorno in FACE_WSDL else "pruebas",
        "proveedor_email": proveedor_email or "",
        "cert_titular": datos["titular"],
        "cert_valido_desde": datos["valido_desde"],
        "cert_valido_hasta": datos["valido_hasta"],
        "cert_emisor": datos["emisor"],
        "cert_subido_at": now_iso(),
    }
    await db.ajustes.update_one({"_id": "config"}, {"$set": {"facturae": fe}}, upsert=True)
    return await obtener_ajustes()


@api_router.put("/facturae/config")
async def guardar_config_facturae(entorno: str = Form(None), proveedor_email: str = Form(None)):
    """Actualiza entorno/email sin volver a subir el certificado."""
    cfg = await _get_ajustes()
    fe = dict(cfg.get("facturae") or {})
    if entorno is not None and entorno in FACE_WSDL:
        fe["entorno"] = entorno
    if proveedor_email is not None:
        fe["proveedor_email"] = proveedor_email
    await db.ajustes.update_one({"_id": "config"}, {"$set": {"facturae": fe}}, upsert=True)
    return await obtener_ajustes()


@api_router.delete("/facturae/certificado")
async def eliminar_certificado_facturae():
    await db.ajustes.update_one({"_id": "config"}, {"$unset": {"facturae": ""}})
    return await obtener_ajustes()


async def _cert_facturae() -> dict:
    cfg = await _get_ajustes()
    fe = cfg.get("facturae") or {}
    if not fe.get("cert_p12_base64"):
        raise HTTPException(400, "No hay certificado configurado. Súbelo en Ajustes → Facturación electrónica.")
    return fe


async def _factura_firmada(doc_id: str) -> tuple:
    """Genera el XML Facturae y lo firma. Devuelve (factura, xml_firmado_bytes, nombre)."""
    f = await db.facturas_emitidas.find_one({"id": doc_id}, {"_id": 0})
    if not f:
        raise HTTPException(404, "Factura no encontrada")
    fe = await _cert_facturae()
    cfg = await _get_ajustes()
    empresa = cfg.get("empresa", {})
    cliente = {}
    if f.get("cliente_id"):
        cliente = await db.contactos.find_one({"id": f["cliente_id"]}, {"_id": 0}) or {}
    if not cliente and f.get("cliente_nif"):
        cliente = await db.contactos.find_one({"tipo": "cliente", "nif": f["cliente_nif"]}, {"_id": 0}) or {}
    if not cliente:
        cliente = {"nombre": f.get("cliente_nombre", ""), "nif": f.get("cliente_nif", "")}
    xml = _facturae_xml(f, empresa, cliente)
    p12 = base64.b64decode(fe["cert_p12_base64"])
    try:
        firmado = firmar_facturae(xml.encode("utf-8"), p12, fe["cert_password"])
    except Exception as e:
        logging.exception("Error firmando Facturae")
        raise HTTPException(500, f"No se pudo firmar la factura: {e}")
    nombre = f"facturae_{(f.get('numero_completo') or doc_id).replace(' ', '_')}.xsig"
    return f, firmado, nombre


@api_router.get("/facturas-emitidas/{doc_id}/facturae-firmado")
async def facturae_firmado(doc_id: str):
    _f, firmado, nombre = await _factura_firmada(doc_id)
    return Response(content=firmado, media_type="application/xml",
                    headers={"Content-Disposition": f'attachment; filename="{nombre}"'})


@api_router.post("/facturas-emitidas/{doc_id}/enviar-face")
async def enviar_factura_face(doc_id: str):
    """Firma la factura y la presenta en FACe (SOAP). Guarda el resultado en la factura."""
    fe = await _cert_facturae()
    f, firmado, _nombre = await _factura_firmada(doc_id)
    if not (f.get("cliente_id") or f.get("cliente_nif")):
        raise HTTPException(400, "La factura no tiene cliente para presentar en FACe.")
    entorno = fe.get("entorno") or "pruebas"
    wsdl = FACE_WSDL[entorno]

    def _enviar():
        import tempfile as _tmp
        from zeep import Client, Settings
        from zeep.transports import Transport
        from zeep.wsse.signature import BinarySignature
        from requests import Session

        class _SignOnly(BinarySignature):
            def verify(self, envelope):
                return envelope  # FACe firma la respuesta con su propio cert; no la verificamos
        p12 = base64.b64decode(fe["cert_p12_base64"])
        from cryptography.hazmat.primitives.serialization import pkcs12, Encoding, PrivateFormat, NoEncryption
        key, cert, _ = pkcs12.load_key_and_certificates(p12, fe["cert_password"].encode())
        keyf = _tmp.NamedTemporaryFile(delete=False, suffix=".pem")
        keyf.write(key.private_bytes(Encoding.PEM, PrivateFormat.PKCS8, NoEncryption()))
        keyf.flush()
        certf = _tmp.NamedTemporaryFile(delete=False, suffix=".pem")
        certf.write(cert.public_bytes(Encoding.PEM))
        certf.flush()
        # TLS cliente (mutuo) + firma WS-Security del sobre SOAP (exigida por FACe)
        combf = _tmp.NamedTemporaryFile(delete=False, suffix=".pem")
        combf.write(cert.public_bytes(Encoding.PEM))
        combf.write(key.private_bytes(Encoding.PEM, PrivateFormat.PKCS8, NoEncryption()))
        combf.flush()
        session = Session()
        session.cert = combf.name
        transport = Transport(session=session, timeout=40)
        wsse = _SignOnly(keyf.name, certf.name)
        client = Client(wsdl, transport=transport, wsse=wsse,
                        settings=Settings(strict=False, xml_huge_tree=True))
        factura_file = {
            "factura": base64.b64encode(firmado).decode(),
            "nombre": _nombre,
            "mime": "text/xml",
        }
        req_type = client.get_type("ns0:EnviarFacturaRequest")
        request = req_type(
            correo=fe.get("proveedor_email") or "",
            factura=factura_file,
            anexos=[],
        )
        return client.service.enviarFactura(request)

    try:
        resultado = await asyncio.get_event_loop().run_in_executor(None, _enviar)
    except Exception as e:
        logging.exception("Error enviando a FACe")
        raise HTTPException(400, f"Error al presentar en FACe ({entorno}): {e}")

    res = getattr(resultado, "resultado", None)
    fac = getattr(resultado, "factura", None)
    codigo = getattr(res, "codigo", "") if res is not None else ""
    descripcion = getattr(res, "descripcion", "") if res is not None else ""
    num_registro = getattr(fac, "numeroRegistro", "") if fac is not None else ""
    face = {
        "presentada_at": now_iso(),
        "entorno": entorno,
        "numero_registro": str(num_registro or ""),
        "codigo": str(codigo or ""),
        "descripcion": str(descripcion or ""),
    }
    await db.facturas_emitidas.update_one({"id": doc_id}, {"$set": {"face": face}})
    return {"ok": True, "face": face}


# ---------------------------------------------------------------------------
# DOCUMENTOS genéricos (Pedidos / Albaranes)
# ---------------------------------------------------------------------------
async def _numero_documento(ambito: str, tipo_doc: str, prefijo: str, serie: str = "") -> str:
    year = datetime.now().year
    if serie:
        n = await _siguiente_contador(ambito, tipo_doc, serie)
        if n is not None:
            return f"{serie}-{year}-{n:04d}"
    n = await _next_seq(f"num_{tipo_doc}")
    return f"{prefijo}-{year}-{n:04d}"


def _build_documento(data: DocumentoInput, numero: str):
    lineas, base, iva, total = calcular_lineas([l.model_dump() for l in data.lineas])
    return {
        "id": new_id(),
        "numero": numero,
        "serie": data.serie,
        "tipo_operacion": data.tipo_operacion,
        "contacto_id": data.contacto_id,
        "contacto_nombre": data.contacto_nombre,
        "contacto_nif": data.contacto_nif,
        "fecha": data.fecha,
        "estado": data.estado,
        "lineas": lineas,
        "base_total": base,
        "iva_total": iva,
        "total": total,
        "pdf_path": data.pdf_path,
        "pdf_filename": data.pdf_filename,
        "vehiculo_id": data.vehiculo_id,
        "vehiculo_matricula": data.vehiculo_matricula,
        "notas": data.notas,
        "created_at": now_iso(),
    }


def _make_documento_routes(entidad: str, coleccion: str, prefijo: str, registrar_entrada=False):
    @api_router.post(f"/{entidad}")
    async def crear(data: DocumentoInput):
        serie = (data.serie or "").strip().upper()
        numero = await _numero_documento(data.tipo_operacion, coleccion, prefijo, serie)
        doc = _build_documento(data, numero)
        if doc["tipo_operacion"] == "venta" and doc["contacto_nombre"] and not doc["contacto_id"]:
            cli = await ensure_cliente(doc["contacto_nombre"], doc.get("contacto_nif", ""))
            if cli:
                doc["contacto_id"] = cli["id"]
                doc["contacto_nif"] = doc["contacto_nif"] or cli.get("nif", "")
        elif doc["tipo_operacion"] == "compra" and doc["contacto_nombre"] and not doc["contacto_id"]:
            prov = await ensure_proveedor(doc["contacto_nombre"], doc.get("contacto_nif", ""))
            if prov:
                doc["contacto_id"] = prov["id"]
                doc["contacto_nif"] = doc["contacto_nif"] or prov.get("nif", "")
        await db[coleccion].insert_one(dict(doc))
        if registrar_entrada and doc["tipo_operacion"] == "compra":
            await registrar_articulos_entrada(doc["lineas"], {
                "tipo": entidad, "numero": doc["numero"], "id": doc["id"],
                "fecha": doc["fecha"], "proveedor": doc["contacto_nombre"],
            })
        return clean(doc)

    @api_router.get(f"/{entidad}")
    async def listar(tipo_operacion: Optional[str] = None):
        q = {}
        if tipo_operacion:
            q['tipo_operacion'] = tipo_operacion
        return await db[coleccion].find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)

    @api_router.get(f"/{entidad}/{{doc_id}}")
    async def obtener(doc_id: str):
        doc = await db[coleccion].find_one({"id": doc_id}, {"_id": 0})
        if not doc:
            raise HTTPException(404, "No encontrado")
        return doc

    @api_router.put(f"/{entidad}/{{doc_id}}")
    async def actualizar(doc_id: str, data: DocumentoInput):
        existing = await db[coleccion].find_one({"id": doc_id})
        if not existing:
            raise HTTPException(404, "No encontrado")
        doc = _build_documento(data, existing['numero'])
        doc['id'] = doc_id
        doc['created_at'] = existing.get('created_at', now_iso())
        await db[coleccion].update_one({"id": doc_id}, {"$set": clean(dict(doc))})
        return doc

    @api_router.delete(f"/{entidad}/{{doc_id}}")
    async def eliminar(doc_id: str):
        await db[coleccion].delete_one({"id": doc_id})
        return {"ok": True}


_make_documento_routes("presupuestos", "presupuestos", "PRE")
_make_documento_routes("pedidos", "pedidos", "PED")
_make_documento_routes("albaranes", "albaranes", "ALB", registrar_entrada=True)


# ---------------------------------------------------------------------------
# FACTURAS EMITIDAS  (con Verifactu)
# ---------------------------------------------------------------------------
async def _generar_verifactu(serie: str, numero: str, fecha: str, nif: str, total: float):
    """Genera la huella encadenada y los datos del QR (compatible Verifactu)."""
    ultima = await db.facturas_emitidas.find_one(
        {"verifactu.huella": {"$exists": True}},
        {"_id": 0, "verifactu": 1},
        sort=[("created_at", -1)],
    )
    huella_anterior = ultima["verifactu"]["huella"] if ultima else ""
    cadena = f"{huella_anterior}|{serie}{numero}|{fecha}|{nif}|{total:.2f}"
    huella = hashlib.sha256(cadena.encode()).hexdigest().upper()
    qr_data = (
        "https://prewww2.aeat.es/wlpl/TIKE-CONT/ValidarQR?"
        f"nif={nif}&numserie={serie}{numero}&fecha={fecha}&importe={total:.2f}"
    )
    return {
        "huella": huella,
        "huella_anterior": huella_anterior,
        "qr_data": qr_data,
        "estado": "registrada",
        "timestamp": now_iso(),
    }


async def _emitir_factura(serie, cliente_id, cliente_nombre, cliente_nif, fecha, lineas, base, iva, total,
                          estado, forma_pago, notas, tipo_factura="ordinaria", rectifica_a=""):
    year = datetime.now().year
    n = await _siguiente_contador("venta", "facturas", serie)
    if n is None:
        n = await db.facturas_emitidas.count_documents({"serie": serie}) + 1
    numero = f"{n:04d}"
    numero_completo = f"{serie}{year}-{numero}"
    verifactu = await _generar_verifactu(serie, numero_completo, fecha, cliente_nif or "B00000000", total)
    doc = {
        "id": new_id(),
        "serie": serie,
        "numero": numero,
        "numero_completo": numero_completo,
        "tipo_factura": tipo_factura,
        "rectifica_a": rectifica_a,
        "cliente_id": cliente_id,
        "cliente_nombre": cliente_nombre,
        "cliente_nif": cliente_nif,
        "fecha_expedicion": fecha,
        "lineas": lineas,
        "base_total": base,
        "iva_total": iva,
        "total": total,
        "estado": estado,
        "forma_pago": forma_pago,
        "verifactu": verifactu,
        "notas": notas,
        "created_at": now_iso(),
    }
    await db.facturas_emitidas.insert_one(dict(doc))
    return clean(doc)


@api_router.post("/facturas-emitidas")
async def crear_factura_emitida(data: FacturaEmitidaInput):
    cliente_id = data.cliente_id
    cliente_nif = data.cliente_nif
    if data.cliente_nombre and not cliente_id:
        cli = await ensure_cliente(data.cliente_nombre, data.cliente_nif)
        if cli:
            cliente_id = cli["id"]
            cliente_nif = cliente_nif or cli.get("nif", "")
    lineas, base, iva, total = calcular_lineas([l.model_dump() for l in data.lineas])
    return await _emitir_factura(data.serie, cliente_id, data.cliente_nombre, cliente_nif,
                                 data.fecha_expedicion, lineas, base, iva, total,
                                 data.estado, data.forma_pago, data.notas)


@api_router.get("/facturas-emitidas")
async def listar_facturas_emitidas():
    return await db.facturas_emitidas.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)


@api_router.get("/facturas-emitidas/{doc_id}")
async def obtener_factura_emitida(doc_id: str):
    doc = await db.facturas_emitidas.find_one({"id": doc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "No encontrada")
    return doc


# ---------------------------------------------------------------------------
# FACTURAE 3.2.2 (factura electrónica para Administraciones Públicas / FACe)
# ---------------------------------------------------------------------------
def _xe(v) -> str:
    return (str(v if v is not None else "")).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def _person_type(nif: str) -> str:
    c = (nif or "").strip().upper()[:1]
    return "J" if c in "ABCDEFGHJNPQRSUVW" else "F"


def _fmt(n, dec=2) -> str:
    return f"{float(n or 0):.{dec}f}"


def _facturae_domicilio(direccion, cp, ciudad):
    return (
        "<AddressInSpain>"
        f"<Address>{_xe(direccion or 'N/A')}</Address>"
        f"<PostCode>{_xe((cp or '00000').strip()[:5].zfill(5))}</PostCode>"
        f"<Town>{_xe(ciudad or 'N/A')}</Town>"
        f"<Province>{_xe(ciudad or 'N/A')}</Province>"
        "<CountryCode>ESP</CountryCode>"
        "</AddressInSpain>"
    )


def _facturae_parte(nif, nombre, direccion, cp, ciudad, centros_xml=""):
    ptype = _person_type(nif)
    entidad = (
        f"<LegalEntity><CorporateName>{_xe(nombre)}</CorporateName>{_facturae_domicilio(direccion, cp, ciudad)}</LegalEntity>"
        if ptype == "J" else
        f"<Individual><Name>{_xe(nombre)}</Name><FirstSurname>{_xe(nombre)}</FirstSurname>{_facturae_domicilio(direccion, cp, ciudad)}</Individual>"
    )
    return (
        "<TaxIdentification>"
        f"<PersonTypeCode>{ptype}</PersonTypeCode>"
        "<ResidenceTypeCode>R</ResidenceTypeCode>"
        f"<TaxIdentificationNumber>{_xe((nif or '').strip().upper())}</TaxIdentificationNumber>"
        "</TaxIdentification>"
        f"{centros_xml}{entidad}"
    )


def _facturae_centros(cliente):
    """AdministrativeCentres con los DIR3 (01 Oficina Contable, 02 Órgano Gestor, 03 Unidad Tramitadora)."""
    mapa = [
        ("01", cliente.get("dir3_oficina_contable", ""), "Oficina contable"),
        ("02", cliente.get("dir3_organo_gestor", ""), "Órgano gestor"),
        ("03", cliente.get("dir3_unidad_tramitadora", ""), "Unidad tramitadora"),
    ]
    centros = ""
    for role, code, nombre in mapa:
        if not code:
            continue
        centros += (
            "<AdministrativeCentre>"
            f"<CentreCode>{_xe(code.strip())}</CentreCode>"
            f"<RoleTypeCode>{role}</RoleTypeCode>"
            f"<Name>{_xe(nombre)}</Name>"
            f"{_facturae_domicilio(cliente.get('direccion'), cliente.get('codigo_postal'), cliente.get('ciudad'))}"
            f"<CentreDescription>{_xe(nombre)}</CentreDescription>"
            "</AdministrativeCentre>"
        )
    return f"<AdministrativeCentres>{centros}</AdministrativeCentres>" if centros else ""


def _facturae_xml(f: dict, empresa: dict, cliente: dict) -> str:
    lineas, base_total, iva_total, total = calcular_lineas(f.get("lineas", []))

    grupos = {}
    for l in lineas:
        g = grupos.setdefault(float(l.get("tipo_iva", 0)), {"base": 0.0, "cuota": 0.0})
        g["base"] += l.get("base", 0)
        g["cuota"] += l.get("cuota_iva", 0)
    taxes = "".join(
        "<Tax><TaxTypeCode>01</TaxTypeCode>"
        f"<TaxRate>{_fmt(rate)}</TaxRate>"
        f"<TaxableBase><TotalAmount>{_fmt(v['base'])}</TotalAmount></TaxableBase>"
        f"<TaxAmount><TotalAmount>{_fmt(v['cuota'])}</TotalAmount></TaxAmount></Tax>"
        for rate, v in sorted(grupos.items(), reverse=True)
    )

    items = ""
    for l in lineas:
        cant = float(l.get("cantidad", 0) or 0)
        precio = float(l.get("precio_unitario", 0) or 0)
        dto = float(l.get("descuento", 0) or 0)
        total_cost = round(cant * precio, 2)
        base = l.get("base", 0)
        descuento_xml = ""
        if dto:
            descuento_xml = (
                "<DiscountsAndRebates><Discount>"
                f"<DiscountReason>Descuento</DiscountReason>"
                f"<DiscountRate>{_fmt(dto)}</DiscountRate>"
                f"<DiscountAmount>{_fmt(round(total_cost - base, 2))}</DiscountAmount>"
                "</Discount></DiscountsAndRebates>"
            )
        items += (
            "<InvoiceLine>"
            f"<ItemDescription>{_xe(l.get('descripcion', ''))}</ItemDescription>"
            f"<Quantity>{_fmt(cant, 6)}</Quantity>"
            "<UnitOfMeasure>01</UnitOfMeasure>"
            f"<UnitPriceWithoutTax>{_fmt(precio, 6)}</UnitPriceWithoutTax>"
            f"<TotalCost>{_fmt(total_cost)}</TotalCost>"
            f"{descuento_xml}"
            f"<GrossAmount>{_fmt(base)}</GrossAmount>"
            "<TaxesOutputs><Tax><TaxTypeCode>01</TaxTypeCode>"
            f"<TaxRate>{_fmt(l.get('tipo_iva', 0))}</TaxRate>"
            f"<TaxableBase><TotalAmount>{_fmt(base)}</TotalAmount></TaxableBase>"
            f"<TaxAmount><TotalAmount>{_fmt(l.get('cuota_iva', 0))}</TotalAmount></TaxAmount>"
            "</Tax></TaxesOutputs>"
            "</InvoiceLine>"
        )

    seller = _facturae_parte(empresa.get("nif"), empresa.get("nombre") or "Emisor",
                             empresa.get("direccion"), empresa.get("codigo_postal"), empresa.get("ciudad"))
    buyer = _facturae_parte(cliente.get("nif") or f.get("cliente_nif"), cliente.get("nombre") or f.get("cliente_nombre") or "Cliente",
                            cliente.get("direccion"), cliente.get("codigo_postal"), cliente.get("ciudad"),
                            centros_xml=_facturae_centros(cliente))

    ns = "http://www.facturae.gob.es/formato/Versiones/Facturaev3_2_2.xml"
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        f'<fe:Facturae xmlns:fe="{ns}">'
        "<FileHeader>"
        "<SchemaVersion>3.2.2</SchemaVersion>"
        "<Modality>I</Modality>"
        "<InvoiceIssuerType>EM</InvoiceIssuerType>"
        "<Batch>"
        f"<BatchIdentifier>{_xe(f.get('numero_completo', ''))}</BatchIdentifier>"
        "<InvoicesCount>1</InvoicesCount>"
        f"<TotalInvoicesAmount><TotalAmount>{_fmt(total)}</TotalAmount></TotalInvoicesAmount>"
        f"<TotalOutstandingAmount><TotalAmount>{_fmt(total)}</TotalAmount></TotalOutstandingAmount>"
        f"<TotalExecutableAmount><TotalAmount>{_fmt(total)}</TotalAmount></TotalExecutableAmount>"
        "<InvoiceCurrencyCode>EUR</InvoiceCurrencyCode>"
        "</Batch>"
        "</FileHeader>"
        f"<Parties><SellerParty>{seller}</SellerParty><BuyerParty>{buyer}</BuyerParty></Parties>"
        "<Invoices><Invoice>"
        "<InvoiceHeader>"
        f"<InvoiceNumber>{_xe(f.get('numero_completo', ''))}</InvoiceNumber>"
        f"<InvoiceSeriesCode>{_xe(f.get('serie', ''))}</InvoiceSeriesCode>"
        "<InvoiceDocumentType>FC</InvoiceDocumentType>"
        f"<InvoiceClass>{'OR' if f.get('tipo_factura') == 'rectificativa' else 'OO'}</InvoiceClass>"
        "</InvoiceHeader>"
        "<InvoiceIssueData>"
        f"<IssueDate>{_xe(f.get('fecha_expedicion', date.today().isoformat()))}</IssueDate>"
        "<InvoiceCurrencyCode>EUR</InvoiceCurrencyCode>"
        "<TaxCurrencyCode>EUR</TaxCurrencyCode>"
        "<LanguageName>es</LanguageName>"
        "</InvoiceIssueData>"
        f"<TaxesOutputs>{taxes}</TaxesOutputs>"
        "<InvoiceTotals>"
        f"<TotalGrossAmount>{_fmt(base_total)}</TotalGrossAmount>"
        f"<TotalGrossAmountBeforeTaxes>{_fmt(base_total)}</TotalGrossAmountBeforeTaxes>"
        f"<TotalTaxOutputs>{_fmt(iva_total)}</TotalTaxOutputs>"
        "<TotalTaxesWithheld>0.00</TotalTaxesWithheld>"
        f"<InvoiceTotal>{_fmt(total)}</InvoiceTotal>"
        f"<TotalOutstandingAmount>{_fmt(total)}</TotalOutstandingAmount>"
        f"<TotalExecutableAmount>{_fmt(total)}</TotalExecutableAmount>"
        "</InvoiceTotals>"
        f"<Items>{items}</Items>"
        "</Invoice></Invoices>"
        "</fe:Facturae>"
    )


@api_router.get("/facturas-emitidas/{doc_id}/facturae")
async def facturae_factura_emitida(doc_id: str):
    f = await db.facturas_emitidas.find_one({"id": doc_id}, {"_id": 0})
    if not f:
        raise HTTPException(404, "No encontrada")
    cfg = await _get_ajustes()
    empresa = cfg.get("empresa", {})
    cliente = {}
    if f.get("cliente_id"):
        cliente = await db.contactos.find_one({"id": f["cliente_id"]}, {"_id": 0}) or {}
    if not cliente and f.get("cliente_nif"):
        cliente = await db.contactos.find_one({"tipo": "cliente", "nif": f["cliente_nif"]}, {"_id": 0}) or {}
    if not cliente:
        cliente = {"nombre": f.get("cliente_nombre", ""), "nif": f.get("cliente_nif", "")}
    xml = _facturae_xml(f, empresa, cliente)
    filename = f"facturae_{(f.get('numero_completo') or doc_id).replace(' ', '_')}.xml"
    return Response(content=xml, media_type="application/xml",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@api_router.patch("/facturas-emitidas/{doc_id}/estado")
async def cambiar_estado_factura_emitida(doc_id: str, estado: str = Form(...)):
    res = await db.facturas_emitidas.update_one({"id": doc_id}, {"$set": {"estado": estado}})
    if res.matched_count == 0:
        raise HTTPException(404, "No encontrada")
    return await db.facturas_emitidas.find_one({"id": doc_id}, {"_id": 0})


@api_router.post("/facturas-emitidas/{doc_id}/rectificar")
async def rectificar_factura_emitida(doc_id: str):
    """Verifactu: las facturas no se eliminan; se emite una factura rectificativa (abono)."""
    orig = await db.facturas_emitidas.find_one({"id": doc_id}, {"_id": 0})
    if not orig:
        raise HTTPException(404, "No encontrada")
    if orig.get("tipo_factura") == "rectificativa":
        raise HTTPException(400, "Una factura rectificativa no se puede rectificar")
    lineas_neg = [{**l, "cantidad": -abs(float(l.get("cantidad", 0)))} for l in orig["lineas"]]
    lineas, base, iva, total = calcular_lineas(lineas_neg)
    nueva = await _emitir_factura(
        orig["serie"], orig.get("cliente_id", ""), orig["cliente_nombre"], orig.get("cliente_nif", ""),
        date.today().isoformat(), lineas, base, iva, total, "emitida",
        orig.get("forma_pago", "Transferencia"), f"Rectificativa (abono) de {orig['numero_completo']}",
        tipo_factura="rectificativa", rectifica_a=orig["numero_completo"],
    )
    await db.facturas_emitidas.update_one({"id": doc_id}, {"$set": {"estado": "rectificada"}})
    return nueva


# ---------------------------------------------------------------------------
# FACTURAS RECIBIDAS
# ---------------------------------------------------------------------------
@api_router.get("/albaranes-compra-pendientes")
async def albaranes_compra_pendientes(proveedor_id: str = "", proveedor_nombre: str = ""):
    """Albaranes de compra aún no facturados (para conciliar con una factura recibida)."""
    q = {"tipo_operacion": "compra", "estado": {"$ne": "facturado"}}
    if proveedor_id:
        q["contacto_id"] = proveedor_id
    elif proveedor_nombre:
        q["contacto_nombre"] = {"$regex": f"^{re.escape(proveedor_nombre)}$", "$options": "i"}
    return await db.albaranes.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)


@api_router.post("/facturas-recibidas")
async def crear_factura_recibida(data: FacturaRecibidaInput):
    lineas, base, iva, total = calcular_lineas([l.model_dump() for l in data.lineas])
    conciliacion = None
    if data.albaranes_ids:
        albs = await db.albaranes.find({"id": {"$in": data.albaranes_ids}}, {"_id": 0}).to_list(500)
        suma = round(sum(a.get("total", 0) for a in albs), 2)
        conciliacion = {
            "albaranes": [{"id": a["id"], "numero": a.get("numero", ""), "total": a.get("total", 0)} for a in albs],
            "suma_albaranes": suma,
            "coincide": abs(suma - total) < 0.01,
        }
    prov_id = data.proveedor_id
    prov_nif = data.proveedor_nif
    if data.proveedor_nombre and not prov_id:
        prov = await ensure_proveedor(data.proveedor_nombre, data.proveedor_nif)
        if prov:
            prov_id = prov["id"]
            prov_nif = prov_nif or prov.get("nif", "")
    doc = {
        "id": new_id(),
        "numero_proveedor": data.numero_proveedor,
        "tipo_factura": "ordinaria",
        "rectifica_a": "",
        "proveedor_id": prov_id,
        "proveedor_nombre": data.proveedor_nombre,
        "proveedor_nif": prov_nif,
        "fecha": data.fecha,
        "lineas": lineas,
        "base_total": base,
        "iva_total": iva,
        "total": total,
        "estado": data.estado,
        "origen": data.origen,
        "forma_pago": data.forma_pago,
        "pdf_base64": data.pdf_base64,
        "pdf_path": data.pdf_path,
        "pdf_filename": data.pdf_filename,
        "albaranes_ids": data.albaranes_ids,
        "conciliacion": conciliacion,
        "vehiculo_id": data.vehiculo_id,
        "vehiculo_matricula": data.vehiculo_matricula,
        "notas": data.notas,
        "created_at": now_iso(),
    }
    await db.facturas_recibidas.insert_one(dict(doc))
    if data.albaranes_ids:
        await db.albaranes.update_many({"id": {"$in": data.albaranes_ids}}, {"$set": {
            "estado": "facturado", "factura_id": doc["id"],
        }})
    await registrar_articulos_entrada(doc["lineas"], {
        "tipo": "factura_recibida", "numero": doc["numero_proveedor"] or doc["id"][:8],
        "id": doc["id"], "fecha": doc["fecha"], "proveedor": doc["proveedor_nombre"],
    })
    return clean(doc)


@api_router.get("/facturas-recibidas")
async def listar_facturas_recibidas():
    docs = await db.facturas_recibidas.find({}, {"_id": 0, "pdf_base64": 0}).sort("created_at", -1).to_list(2000)
    return docs


@api_router.get("/facturas-recibidas/{doc_id}")
async def obtener_factura_recibida(doc_id: str):
    doc = await db.facturas_recibidas.find_one({"id": doc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "No encontrada")
    return doc


@api_router.patch("/facturas-recibidas/{doc_id}/estado")
async def cambiar_estado_factura_recibida(doc_id: str, estado: str = Form(...)):
    res = await db.facturas_recibidas.update_one({"id": doc_id}, {"$set": {"estado": estado}})
    if res.matched_count == 0:
        raise HTTPException(404, "No encontrada")
    return await db.facturas_recibidas.find_one({"id": doc_id}, {"_id": 0, "pdf_base64": 0})


@api_router.post("/facturas-recibidas/{doc_id}/rectificar")
async def rectificar_factura_recibida(doc_id: str):
    """Verifactu: no se elimina; se registra una rectificativa (abono) que referencia la original."""
    orig = await db.facturas_recibidas.find_one({"id": doc_id}, {"_id": 0})
    if not orig:
        raise HTTPException(404, "No encontrada")
    if orig.get("tipo_factura") == "rectificativa":
        raise HTTPException(400, "Una factura rectificativa no se puede rectificar")
    lineas_neg = [{**l, "cantidad": -abs(float(l.get("cantidad", 0)))} for l in orig["lineas"]]
    lineas, base, iva, total = calcular_lineas(lineas_neg)
    doc = {
        "id": new_id(),
        "numero_proveedor": f"RECT-{orig.get('numero_proveedor', '')}",
        "tipo_factura": "rectificativa",
        "rectifica_a": orig.get("numero_proveedor", "") or orig["id"][:8],
        "proveedor_id": orig.get("proveedor_id", ""),
        "proveedor_nombre": orig["proveedor_nombre"],
        "proveedor_nif": orig.get("proveedor_nif", ""),
        "fecha": date.today().isoformat(),
        "lineas": lineas,
        "base_total": base,
        "iva_total": iva,
        "total": total,
        "estado": "pendiente",
        "origen": "rectificativa",
        "forma_pago": orig.get("forma_pago", "Transferencia"),
        "pdf_base64": "",
        "notas": f"Rectificativa (abono) de {orig.get('numero_proveedor', '')}",
        "created_at": now_iso(),
    }
    await db.facturas_recibidas.insert_one(dict(doc))
    await db.facturas_recibidas.update_one({"id": doc_id}, {"$set": {"estado": "rectificada"}})
    return clean(doc)


# ---------------------------------------------------------------------------
# CONVERSIÓN entre documentos (presupuesto → pedido/albarán → factura)
# ---------------------------------------------------------------------------
_PREFIJOS = {"presupuestos": "PRE", "pedidos": "PED", "albaranes": "ALB"}
_TRANSICIONES = {
    "presupuestos": ["pedidos", "albaranes"],
    "pedidos": ["albaranes"],
    "albaranes": ["factura"],
}


class ConvertirInput(BaseModel):
    destino: str  # "pedidos" | "albaranes" | "factura"


@api_router.post("/documentos/{entidad}/{doc_id}/convertir")
async def convertir_documento(entidad: str, doc_id: str, data: ConvertirInput):
    if entidad not in _TRANSICIONES:
        raise HTTPException(400, "Origen no válido")
    if data.destino not in _TRANSICIONES[entidad]:
        raise HTTPException(400, f"No se puede convertir {entidad} a {data.destino}")
    src = await db[entidad].find_one({"id": doc_id}, {"_id": 0})
    if not src:
        raise HTTPException(404, "Documento no encontrado")
    if src.get("convertido_a") == data.destino or src.get("estado") == "facturado":
        raise HTTPException(400, "Este documento ya se ha convertido")
    op = src.get("tipo_operacion", "venta")
    lineas, base, iva, total = calcular_lineas(src.get("lineas", []))
    serie = src.get("serie", "")

    if data.destino in ("pedidos", "albaranes"):
        numero = await _numero_documento(op, data.destino, _PREFIJOS[data.destino], serie)
        nuevo = {
            "id": new_id(), "numero": numero, "serie": serie, "tipo_operacion": op,
            "contacto_id": src.get("contacto_id", ""), "contacto_nombre": src.get("contacto_nombre", ""),
            "contacto_nif": src.get("contacto_nif", ""), "fecha": date.today().isoformat(),
            "estado": "confirmado", "lineas": lineas, "base_total": base, "iva_total": iva, "total": total,
            "notas": f"Generado desde {entidad[:-1]} {src.get('numero', '')}",
            "origen_tipo": entidad, "origen_id": doc_id, "origen_numero": src.get("numero", ""),
            "created_at": now_iso(),
        }
        await db[data.destino].insert_one(dict(nuevo))
        if data.destino == "albaranes" and op == "compra":
            await registrar_articulos_entrada(lineas, {
                "tipo": "albaranes", "numero": numero, "id": nuevo["id"],
                "fecha": nuevo["fecha"], "proveedor": nuevo["contacto_nombre"],
            })
        await db[entidad].update_one({"id": doc_id}, {"$set": {
            "convertido_a": data.destino, "convertido_ref": numero, "estado": "confirmado",
        }})
        return clean(nuevo)

    # destino == "factura"
    if op == "venta":
        cfg = await _get_ajustes()
        sv = cfg.get("series_venta", [])
        if not any(s["nombre"] == serie for s in sv):
            serie = (next((s for s in sv if s.get("por_defecto")), sv[0])["nombre"] if sv else "A")
        fact = await _emitir_factura(
            serie, src.get("contacto_id", ""), src.get("contacto_nombre", ""), src.get("contacto_nif", ""),
            date.today().isoformat(), lineas, base, iva, total, "emitida", "Transferencia",
            f"Generada desde albarán {src.get('numero', '')}",
        )
        await db[entidad].update_one({"id": doc_id}, {"$set": {
            "estado": "facturado", "factura_id": fact["id"], "factura_numero": fact["numero_completo"],
        }})
        return {"tipo": "emitida", **fact}

    # compra → factura recibida
    prov_id = src.get("contacto_id", "")
    if src.get("contacto_nombre") and not prov_id:
        prov = await ensure_proveedor(src.get("contacto_nombre", ""), src.get("contacto_nif", ""))
        if prov:
            prov_id = prov["id"]
    doc = {
        "id": new_id(), "numero_proveedor": "", "tipo_factura": "ordinaria", "rectifica_a": "",
        "proveedor_id": prov_id, "proveedor_nombre": src.get("contacto_nombre", ""),
        "proveedor_nif": src.get("contacto_nif", ""), "fecha": date.today().isoformat(),
        "lineas": lineas, "base_total": base, "iva_total": iva, "total": total,
        "estado": "pendiente", "origen": "albaran", "forma_pago": "Transferencia", "pdf_base64": "",
        "albaranes_ref": [src.get("numero", "")], "albaranes_ids": [doc_id],
        "notas": f"Generada desde albarán {src.get('numero', '')}", "created_at": now_iso(),
    }
    await db.facturas_recibidas.insert_one(dict(doc))
    await db[entidad].update_one({"id": doc_id}, {"$set": {
        "estado": "facturado", "factura_id": doc["id"], "factura_numero": doc["numero_proveedor"] or doc["id"][:8],
    }})
    return {"tipo": "recibida", **clean(doc)}


# ---------------------------------------------------------------------------
# EXTRACCIÓN IA de PDF
# ---------------------------------------------------------------------------
EXTRACCION_PROMPT = """Eres un experto en contabilidad española. Analiza el documento PDF adjunto \
(que puede ser una factura, un albarán o un pedido de un proveedor) y extrae los datos en formato JSON.

Devuelve EXCLUSIVAMENTE un objeto JSON válido, sin texto adicional ni markdown, con esta estructura:
{
  "tipo_documento": "factura | albaran | pedido",
  "proveedor": {"nombre": "", "nif": "", "direccion": "", "email": "", "telefono": ""},
  "numero": "número del documento",
  "fecha": "YYYY-MM-DD",
  "lineas": [
    {"descripcion": "", "cantidad": 0, "precio_unitario": 0, "descuento": 0, "tipo_iva": 21}
  ],
  "base_total": 0,
  "iva_total": 0,
  "total": 0,
  "moneda": "EUR"
}

Reglas:
- El NIF/CIF español tiene formato como B12345678 o 12345678Z.
- precio_unitario es el precio sin IVA por unidad.
- tipo_iva es el porcentaje de IVA (21, 10, 4 o 0).
- Si un campo no aparece, usa "" para textos y 0 para números.
- Usa punto decimal, nunca coma."""


def _parse_json(text: str) -> dict:
    text = text.strip()
    if text.startswith("```"):
        text = text.split("```", 2)[1]
        if text.startswith("json"):
            text = text[4:]
        text = text.strip().rstrip("`").strip()
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1:
        text = text[start:end + 1]
    return json.loads(text)


@api_router.post("/archivos/subir")
async def subir_archivo(file: UploadFile = File(...)):
    contenido = await file.read()
    if not (file.filename or "").lower().endswith(".pdf"):
        raise HTTPException(400, "Solo se admiten archivos PDF")
    if len(contenido) > 15 * 1024 * 1024:
        raise HTTPException(400, "El PDF no debe superar 15 MB")
    try:
        return await _guardar_pdf(contenido, file.filename)
    except Exception as e:
        logger.exception("Error subiendo PDF")
        raise HTTPException(500, f"No se pudo guardar el archivo: {e}")


@api_router.get("/archivos/{path:path}")
async def obtener_archivo(path: str):
    try:
        data, _ = storage_get(path)
    except Exception:
        raise HTTPException(404, "Archivo no encontrado")
    return Response(content=data, media_type="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="{path.split("/")[-1]}"'})


@api_router.post("/extraccion/pdf")
async def extraer_pdf(file: UploadFile = File(...), licencia: str = Form("")):
    if not EMERGENT_LLM_KEY:
        raise HTTPException(500, "Falta EMERGENT_LLM_KEY")
    contenido = await file.read()
    suffix = ".pdf" if (file.filename or "").lower().endswith(".pdf") else ""
    tmp_path = None
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage, FileContentWithMimeType
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(contenido)
            tmp_path = tmp.name
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"extraccion-{new_id()}",
            system_message="Eres un asistente experto en extraer datos estructurados de documentos comerciales.",
        ).with_model("gemini", "gemini-2.5-flash")
        pdf_file = FileContentWithMimeType(file_path=tmp_path, mime_type="application/pdf")
        resp = await chat.send_message_with_tools(UserMessage(text=EXTRACCION_PROMPT, file_contents=[pdf_file]))
        respuesta = resp.content or ""
        datos = _parse_json(respuesta if isinstance(respuesta, str) else str(respuesta))
        # normaliza líneas
        lineas_raw = datos.get("lineas", []) or []
        lineas, base, iva, total = calcular_lineas(lineas_raw)
        datos["lineas"] = lineas
        datos["base_total"] = base
        datos["iva_total"] = iva
        datos["total"] = total if not datos.get("total") else datos.get("total")
        # intenta emparejar proveedor por NIF
        nif = (datos.get("proveedor") or {}).get("nif", "")
        if nif:
            match = await db.contactos.find_one({"tipo": "proveedor", "nif": nif}, {"_id": 0})
            if match:
                datos["proveedor_existente"] = match
        # guarda el PDF original para vista previa / gestión documental
        try:
            datos.update(await _guardar_pdf(contenido, file.filename))
        except Exception:
            logger.warning("No se pudo guardar el PDF original en object storage")
        # registro de consumo de IA (tokens + coste estimado)
        u = resp.usage
        input_tokens = getattr(u, "input_tokens", 0) or 0
        output_tokens = getattr(u, "output_tokens", 0) or 0
        total_tokens = getattr(u, "total_tokens", input_tokens + output_tokens) or (input_tokens + output_tokens)
        coste_eur = round((input_tokens * PRECIO_INPUT_TOKEN + output_tokens * PRECIO_OUTPUT_TOKEN) * USD_EUR, 5)
        consumo = {
            "id": new_id(),
            "tipo": "extraccion_pdf",
            "license_key": licencia or "",
            "archivo": file.filename or "documento.pdf",
            "modelo": "gemini-2.5-flash",
            "input_tokens": input_tokens,
            "output_tokens": output_tokens,
            "total_tokens": total_tokens,
            "coste_eur": coste_eur,
            "created_at": now_iso(),
        }
        await db.consumos_ia.insert_one(dict(consumo))
        return {"ok": True, "datos": datos, "consumo": clean(consumo)}
    except json.JSONDecodeError:
        raise HTTPException(422, "No se pudo interpretar la respuesta de la IA")
    except Exception as e:
        logger.exception("Error en extracción PDF")
        raise HTTPException(500, f"Error al procesar el PDF: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.remove(tmp_path)


@api_router.get("/consumos-ia/resumen")
async def consumos_ia_resumen():
    docs = await db.consumos_ia.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    total_tokens = sum(d.get("total_tokens", 0) for d in docs)
    coste_total = round(sum(d.get("coste_eur", 0) for d in docs), 4)
    return {
        "num_lecturas": len(docs),
        "total_tokens": total_tokens,
        "coste_total_eur": coste_total,
        "ultimas": docs[:10],
    }


@api_router.get("/admin/consumos-ia")
async def admin_consumos_ia(admin: dict = Depends(get_current_admin)):
    """Consumo de IA agregado por cliente/licencia (panel central)."""
    docs = await db.consumos_ia.find({}, {"_id": 0}).sort("created_at", -1).to_list(20000)
    licencias = await db.licencias.find({}, {"_id": 0, "license_key": 1, "empresa": 1}).to_list(5000)
    empresa_por_clave = {l["license_key"]: l.get("empresa", "") for l in licencias}
    por_cliente = {}
    for d in docs:
        key = d.get("license_key") or "__sin_licencia__"
        agg = por_cliente.setdefault(key, {"license_key": d.get("license_key") or "",
                                           "empresa": empresa_por_clave.get(d.get("license_key"), "Sin identificar"),
                                           "num_lecturas": 0, "total_tokens": 0, "coste_total_eur": 0.0})
        agg["num_lecturas"] += 1
        agg["total_tokens"] += d.get("total_tokens", 0)
        agg["coste_total_eur"] += d.get("coste_eur", 0)
    clientes = sorted(por_cliente.values(), key=lambda x: x["coste_total_eur"], reverse=True)
    for c in clientes:
        c["coste_total_eur"] = round(c["coste_total_eur"], 4)
    return {
        "num_lecturas": len(docs),
        "total_tokens": sum(d.get("total_tokens", 0) for d in docs),
        "coste_total_eur": round(sum(d.get("coste_eur", 0) for d in docs), 4),
        "clientes": clientes,
    }


# ---------------------------------------------------------------------------
# DASHBOARD
# ---------------------------------------------------------------------------
@api_router.get("/dashboard/resumen")
async def dashboard_resumen():
    clientes = await db.contactos.count_documents({"tipo": "cliente"})
    proveedores = await db.contactos.count_documents({"tipo": "proveedor"})
    pedidos = await db.pedidos.count_documents({})
    albaranes = await db.albaranes.count_documents({})
    articulos = await db.articulos.count_documents({})

    emitidas = await db.facturas_emitidas.find({}, {"_id": 0}).to_list(5000)
    recibidas = await db.facturas_recibidas.find({}, {"_id": 0, "pdf_base64": 0}).to_list(5000)

    total_facturado = round(sum(f.get("total", 0) for f in emitidas), 2)
    pendiente_cobro = round(sum(f.get("total", 0) for f in emitidas if f.get("estado") != "cobrada"), 2)
    total_gastos = round(sum(f.get("total", 0) for f in recibidas), 2)
    pendiente_pago = round(sum(f.get("total", 0) for f in recibidas if f.get("estado") != "pagada"), 2)

    # facturación mensual (últimos 6 meses)
    meses = {}
    for f in emitidas:
        mes = (f.get("fecha_expedicion") or "")[:7]
        if mes:
            meses[mes] = meses.get(mes, 0) + f.get("total", 0)
    grafico = [{"mes": k, "total": round(v, 2)} for k, v in sorted(meses.items())][-6:]

    ultimas_emitidas = sorted(emitidas, key=lambda x: x.get("created_at", ""), reverse=True)[:5]
    ultimas_recibidas = sorted(recibidas, key=lambda x: x.get("created_at", ""), reverse=True)[:5]

    return {
        "clientes": clientes,
        "proveedores": proveedores,
        "pedidos": pedidos,
        "albaranes": albaranes,
        "articulos": articulos,
        "num_facturas_emitidas": len(emitidas),
        "num_facturas_recibidas": len(recibidas),
        "total_facturado": total_facturado,
        "pendiente_cobro": pendiente_cobro,
        "total_gastos": total_gastos,
        "pendiente_pago": pendiente_pago,
        "grafico_facturacion": grafico,
        "ultimas_emitidas": ultimas_emitidas,
        "ultimas_recibidas": ultimas_recibidas,
    }

# ---------------------------------------------------------------------------
# MÓDULO TALLER (chapa, pintura y mecánica) — Fase 1: Vehículos + Órdenes
# ---------------------------------------------------------------------------
class VehiculoInput(BaseModel):
    matricula: str = ""
    marca: str = ""
    modelo: str = ""
    bastidor: str = ""
    color: str = ""
    kilometros: Optional[int] = None
    combustible: str = ""
    anio: Optional[int] = None
    cliente_id: str = ""
    cliente_nombre: str = ""
    tipo: Literal['cliente', 'cortesia'] = 'cliente'
    notas: str = ""


class Vehiculo(VehiculoInput):
    id: str = Field(default_factory=new_id)
    propietarios: List[dict] = []
    created_at: str = Field(default_factory=now_iso)


async def _normaliza_vehiculo(d: dict) -> dict:
    if d.get("matricula"):
        d["matricula"] = d["matricula"].upper().strip()
    if d.get("cliente_id"):
        cli = await db.contactos.find_one({"id": d["cliente_id"]}, {"_id": 0})
        if cli:
            d["cliente_nombre"] = cli.get("nombre", "")
    return d


@api_router.post("/taller/vehiculos")
async def crear_vehiculo(data: VehiculoInput):
    d = await _normaliza_vehiculo(data.model_dump())
    mat = d.get("matricula", "")
    if mat:
        ex = await db.vehiculos.find_one({"matricula": mat}, {"_id": 0})
        if ex:
            raise HTTPException(409, f"Ya existe un vehículo con la matrícula {mat}. La matrícula debe ser única.")
    d["propietarios"] = []
    if d.get("cliente_id"):
        d["propietarios"] = [{"cliente_id": d["cliente_id"], "cliente_nombre": d.get("cliente_nombre", ""), "desde": now_iso(), "hasta": None}]
    v = Vehiculo(**d)
    await db.vehiculos.insert_one(v.model_dump())
    return v.model_dump()


@api_router.get("/taller/vehiculos")
async def listar_vehiculos(q: Optional[str] = None):
    docs = await db.vehiculos.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)
    if q:
        ql = q.lower()
        docs = [d for d in docs if ql in " ".join([
            d.get("matricula", ""), d.get("marca", ""), d.get("modelo", ""),
            d.get("bastidor", ""), d.get("cliente_nombre", "")]).lower()]
    return docs


@api_router.get("/taller/vehiculos/{vid}")
async def obtener_vehiculo(vid: str):
    doc = await db.vehiculos.find_one({"id": vid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Vehículo no encontrado")
    return doc


@api_router.get("/taller/vehiculos/{vid}/ficha")
async def ficha_vehiculo(vid: str):
    v = await db.vehiculos.find_one({"id": vid}, {"_id": 0})
    if not v:
        raise HTTPException(404, "Vehículo no encontrado")
    ordenes = await db.ordenes_trabajo.find({"vehiculo_id": vid}, {"_id": 0}).sort("created_at", -1).to_list(500)
    peritajes = await db.peritajes.find({"vehiculo_id": vid}, {"_id": 0}).sort("created_at", -1).to_list(500)
    presupuestos = await db.presupuestos.find({"vehiculo_id": vid}, {"_id": 0}).sort("created_at", -1).to_list(500)
    prestamos = await db.prestamos.find({"vehiculo_id": vid}, {"_id": 0}).sort("created_at", -1).to_list(500)
    citas = await db.citas.find({"vehiculo_id": vid}, {"_id": 0}).sort("fecha", -1).to_list(500)
    compras = []
    for col, etiqueta in [("pedidos", "Pedido"), ("albaranes", "Albarán"), ("facturas_recibidas", "Factura")]:
        docs = await db[col].find({"vehiculo_id": vid}, {"_id": 0}).sort("created_at", -1).to_list(500)
        for d in docs:
            compras.append({
                "tipo": etiqueta,
                "id": d.get("id"),
                "numero": d.get("numero") or d.get("numero_proveedor") or "",
                "fecha": d.get("fecha", ""),
                "proveedor": d.get("contacto_nombre") or d.get("proveedor_nombre") or "",
                "total": d.get("total", 0),
            })
    coste_compras = round(sum(c["total"] for c in compras), 2)
    facturas = await db.facturas_emitidas.find({"vehiculo_id": vid}, {"_id": 0}).sort("created_at", -1).to_list(500)
    # Material y mano de obra asignados al vehículo (líneas de las órdenes de trabajo)
    materiales = []
    for o in ordenes:
        for l in (o.get("lineas") or []):
            materiales.append({
                "orden_id": o.get("id"),
                "orden_numero": o.get("numero", ""),
                "fecha": o.get("fecha_entrada") or o.get("created_at", ""),
                "descripcion": l.get("descripcion", ""),
                "cantidad": l.get("cantidad"),
                "unidad": l.get("unidad", ""),
                "articulo_id": l.get("articulo_id") or l.get("articulo") or None,
                "es_material": bool(l.get("articulo_id")) or (l.get("tipo") in ("material", "articulo")),
                "precio_unitario": l.get("precio_unitario"),
                "total": l.get("total"),
            })
    return {"vehiculo": v, "propietarios": v.get("propietarios", []), "ordenes": ordenes,
            "peritajes": peritajes, "presupuestos": presupuestos, "prestamos": prestamos,
            "citas": citas, "compras": compras, "coste_compras": coste_compras,
            "facturas": facturas, "materiales": materiales}


@api_router.put("/taller/vehiculos/{vid}")
async def actualizar_vehiculo(vid: str, data: VehiculoInput):
    d = await _normaliza_vehiculo(data.model_dump())
    cur = await db.vehiculos.find_one({"id": vid}, {"_id": 0})
    if not cur:
        raise HTTPException(404, "Vehículo no encontrado")
    mat = d.get("matricula", "")
    if mat:
        other = await db.vehiculos.find_one({"matricula": mat, "id": {"$ne": vid}}, {"_id": 0})
        if other:
            raise HTTPException(409, f"Ya existe otro vehículo con la matrícula {mat}. La matrícula debe ser única.")
    props = list(cur.get("propietarios") or [])
    nuevo = d.get("cliente_id")
    anterior = cur.get("cliente_id")
    if nuevo and nuevo != anterior:
        if props and props[-1].get("hasta") is None:
            props[-1]["hasta"] = now_iso()
        props.append({"cliente_id": nuevo, "cliente_nombre": d.get("cliente_nombre", ""), "desde": now_iso(), "hasta": None})
    elif nuevo and not props:
        props = [{"cliente_id": nuevo, "cliente_nombre": d.get("cliente_nombre", ""), "desde": cur.get("created_at", now_iso()), "hasta": None}]
    d["propietarios"] = props
    await db.vehiculos.update_one({"id": vid}, {"$set": d})
    return await db.vehiculos.find_one({"id": vid}, {"_id": 0})


@api_router.delete("/taller/vehiculos/{vid}")
async def eliminar_vehiculo(vid: str):
    await db.vehiculos.delete_one({"id": vid})
    return {"ok": True}


# ---- Órdenes de trabajo ----
ESTADOS_OT = ['recepcion', 'en_curso', 'finalizado', 'entregado']


class OrdenTrabajoInput(BaseModel):
    vehiculo_id: str = ""
    vehiculo_matricula: str = ""
    cliente_id: str = ""
    cliente_nombre: str = ""
    tipos_trabajo: List[str] = []   # chapa / pintura / mecanica
    descripcion: str = ""
    estado: Literal['recepcion', 'en_curso', 'finalizado', 'entregado'] = 'recepcion'
    fecha_entrada: str = ""
    fecha_entrega_estimada: str = ""
    lineas: List[LineaItem] = []
    notas: str = ""


class OrdenTrabajo(OrdenTrabajoInput):
    id: str = Field(default_factory=new_id)
    numero: str = ""
    base: float = 0
    cuota_iva: float = 0
    total: float = 0
    created_at: str = Field(default_factory=now_iso)


async def _rellena_orden(d: dict) -> dict:
    if d.get("vehiculo_id"):
        v = await db.vehiculos.find_one({"id": d["vehiculo_id"]}, {"_id": 0})
        if v:
            d["vehiculo_matricula"] = v.get("matricula", "")
            if not d.get("cliente_id"):
                d["cliente_id"] = v.get("cliente_id", "")
                d["cliente_nombre"] = v.get("cliente_nombre", "")
    if d.get("cliente_id") and not d.get("cliente_nombre"):
        cli = await db.contactos.find_one({"id": d["cliente_id"]}, {"_id": 0})
        if cli:
            d["cliente_nombre"] = cli.get("nombre", "")
    lineas, base, iva, total = calcular_lineas(d.get("lineas", []))
    d["lineas"] = lineas
    d["base"], d["cuota_iva"], d["total"] = base, iva, total
    return d


@api_router.post("/taller/ordenes")
async def crear_orden(data: OrdenTrabajoInput):
    d = await _rellena_orden(data.model_dump())
    n = await _next_seq("orden_trabajo")
    d["numero"] = f"OT-{n:06d}"
    o = OrdenTrabajo(**d)
    await db.ordenes_trabajo.insert_one(o.model_dump())
    return o.model_dump()


@api_router.get("/taller/ordenes")
async def listar_ordenes(vehiculo_id: Optional[str] = None):
    q = {}
    if vehiculo_id:
        q["vehiculo_id"] = vehiculo_id
    return await db.ordenes_trabajo.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)


@api_router.get("/taller/ordenes/{oid}")
async def obtener_orden(oid: str):
    doc = await db.ordenes_trabajo.find_one({"id": oid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Orden no encontrada")
    return doc


@api_router.put("/taller/ordenes/{oid}")
async def actualizar_orden(oid: str, data: OrdenTrabajoInput):
    existing = await db.ordenes_trabajo.find_one({"id": oid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Orden no encontrada")
    d = await _rellena_orden(data.model_dump())
    d["numero"] = existing.get("numero", "")
    await db.ordenes_trabajo.update_one({"id": oid}, {"$set": d})
    return await db.ordenes_trabajo.find_one({"id": oid}, {"_id": 0})


@api_router.patch("/taller/ordenes/{oid}/estado")
async def estado_orden(oid: str, estado: str = Form(...)):
    if estado not in ESTADOS_OT:
        raise HTTPException(400, "Estado inválido")
    res = await db.ordenes_trabajo.update_one({"id": oid}, {"$set": {"estado": estado}})
    if res.matched_count == 0:
        raise HTTPException(404, "Orden no encontrada")
    return await db.ordenes_trabajo.find_one({"id": oid}, {"_id": 0})


@api_router.delete("/taller/ordenes/{oid}")
async def eliminar_orden(oid: str):
    await db.ordenes_trabajo.delete_one({"id": oid})
    return {"ok": True}


# ---------------------------------------------------------------------------
# TALLER — FASE 2: Peritajes, Compañías de seguros, Fotos (subida + QR)
# ---------------------------------------------------------------------------
_TALLER_COLS = {"peritajes": "peritajes", "vehiculos": "vehiculos", "ordenes": "ordenes_trabajo"}


async def _guardar_media(contenido: bytes, filename: str, content_type: str) -> dict:
    ext = filename.rsplit(".", 1)[-1].lower() if filename and "." in filename else "bin"
    path = f"{APP_NAME}/media/{new_id()}.{ext}"
    res = storage_put(path, contenido, content_type or "application/octet-stream")
    return {"path": res.get("path", path), "filename": filename or "archivo", "content_type": content_type or ""}


async def _append_foto(col: str, eid: str, file: UploadFile) -> dict:
    contenido = await file.read()
    ct = file.content_type or ""
    if not (ct.startswith("image/") or ct == "application/pdf"):
        raise HTTPException(400, "Solo se admiten imágenes o PDF")
    if len(contenido) > 15 * 1024 * 1024:
        raise HTTPException(400, "El archivo no debe superar 15 MB")
    media = await _guardar_media(contenido, file.filename, ct)
    foto = {"path": media["path"], "filename": media["filename"], "content_type": ct, "created_at": now_iso()}
    res = await db[col].update_one({"id": eid}, {"$push": {"fotos": foto}})
    if res.matched_count == 0:
        raise HTTPException(404, "Registro no encontrado")
    return foto


@api_router.get("/taller/media/{path:path}")
async def obtener_media(path: str):
    try:
        data, ct = storage_get(path)
    except Exception:
        raise HTTPException(404, "Archivo no encontrado")
    return Response(content=data, media_type=ct or "application/octet-stream",
                    headers={"Content-Disposition": f'inline; filename="{path.split("/")[-1]}"'})


@api_router.post("/taller/{tipo}/{eid}/fotos")
async def subir_foto(tipo: str, eid: str, file: UploadFile = File(...)):
    col = _TALLER_COLS.get(tipo)
    if not col:
        raise HTTPException(404, "Tipo no válido")
    return await _append_foto(col, eid, file)


@api_router.delete("/taller/{tipo}/{eid}/fotos")
async def borrar_foto(tipo: str, eid: str, path: str):
    col = _TALLER_COLS.get(tipo)
    if not col:
        raise HTTPException(404, "Tipo no válido")
    await db[col].update_one({"id": eid}, {"$pull": {"fotos": {"path": path}}})
    return {"ok": True}


class FirmaInput(BaseModel):
    imagen: str  # dataURL base64 (data:image/png;base64,...)


@api_router.post("/taller/ordenes/{oid}/firma")
async def guardar_firma_orden(oid: str, data: FirmaInput):
    m = re.match(r"data:(image/\w+);base64,(.+)", data.imagen or "", re.DOTALL)
    if not m:
        raise HTTPException(400, "Imagen de firma no válida")
    ct, raw = m.group(1), base64.b64decode(m.group(2))
    if len(raw) > 3 * 1024 * 1024:
        raise HTTPException(400, "La firma es demasiado grande")
    media = await _guardar_media(raw, "firma.png", ct)
    ts = now_iso()
    res = await db.ordenes_trabajo.update_one(
        {"id": oid}, {"$set": {"firma_cliente_path": media["path"], "firma_cliente_at": ts}})
    if res.matched_count == 0:
        raise HTTPException(404, "Orden no encontrada")
    return {"firma_cliente_path": media["path"], "firma_cliente_at": ts}


@api_router.delete("/taller/ordenes/{oid}/firma")
async def borrar_firma_orden(oid: str):
    await db.ordenes_trabajo.update_one(
        {"id": oid}, {"$set": {"firma_cliente_path": "", "firma_cliente_at": ""}})
    return {"ok": True}


_TIPOS_TRABAJO = {"chapa": "Chapa", "pintura": "Pintura", "mecanica": "Mecánica",
                  "electricidad": "Electricidad", "diagnosis": "Diagnosis",
                  "revision": "Revisión", "neumaticos": "Neumáticos", "otros": "Otros"}


def _img_data_uri(path_or_url: str) -> str:
    """Descarga logo/firma y lo devuelve como data URI (para incrustar en el PDF)."""
    try:
        if not path_or_url:
            return ""
        if path_or_url.startswith("http"):
            import requests
            r = requests.get(path_or_url, timeout=8)
            if r.status_code != 200:
                return ""
            ct = r.headers.get("content-type", "image/png").split(";")[0]
            return f"data:{ct};base64,{base64.b64encode(r.content).decode()}"
        data, ct = storage_get(path_or_url)
        return f"data:{ct or 'image/png'};base64,{base64.b64encode(data).decode()}"
    except Exception:
        return ""


def _build_resguardo_html(orden, vehiculo, cliente, empresa):
    def e(v):
        return str(v if v is not None else "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def fmtf(f):
        if not f:
            return ""
        try:
            return datetime.fromisoformat(str(f)[:19]).strftime("%d/%m/%Y")
        except Exception:
            return str(f)

    logo = _img_data_uri(empresa.get("logo", ""))
    firma = _img_data_uri(orden.get("firma_cliente_path", ""))
    dir_taller = " · ".join([x for x in [empresa.get("direccion", ""),
                             " ".join([x for x in [empresa.get("codigo_postal", ""), empresa.get("ciudad", "")] if x])] if x])
    trabajos = ", ".join([_TIPOS_TRABAJO.get(t, t) for t in (orden.get("tipos_trabajo") or [])])
    obs = " · ".join([x for x in [orden.get("descripcion", ""), trabajos] if x])

    filas = ""
    lineas = orden.get("lineas") or []
    for i in range(21):
        l = lineas[i] if i < len(lineas) else None
        desc = mo = mat = ""
        if l:
            desc = e(l.get("descripcion", ""))
            imp = float(l.get("cantidad", 0) or 0) * float(l.get("precio_unitario", 0) or 0)
            val = f"{imp:.2f} €"
            if (l.get("unidad") or "").lower() in ("h", "hora", "horas", "mo"):
                mo = val
            else:
                mat = val
        filas += (f'<tr style="height:26px"><td class="n">{i+1}</td>'
                  f'<td class="d">{desc}</td><td class="m">{mo}</td><td class="m">{mat}</td></tr>')

    car = ('<svg viewBox="0 0 300 150" width="150" height="72">'
           '<rect x="46" y="6" width="40" height="15" rx="6" fill="#e4e4e7"/>'
           '<rect x="214" y="6" width="40" height="15" rx="6" fill="#e4e4e7"/>'
           '<rect x="46" y="129" width="40" height="15" rx="6" fill="#e4e4e7"/>'
           '<rect x="214" y="129" width="40" height="15" rx="6" fill="#e4e4e7"/>'
           '<rect x="34" y="18" width="232" height="114" rx="46" fill="none" stroke="#555" stroke-width="1.6"/>'
           '<rect x="102" y="60" width="96" height="30" rx="4" fill="none" stroke="#bbb" stroke-width="1"/>'
           '</svg>')
    logo_html = f'<img src="{logo}" style="max-height:42px;max-width:150px"/>' if logo else ""
    firma_html = f'<img src="{firma}" style="max-height:36px;max-width:150px"/>' if firma else '<div style="height:36px"></div>'

    def cell(lbl, val):
        return f'<td class="c"><div class="lbl">{e(lbl)}</div><div class="val">{e(val)}</div></td>'

    body = f'''
    <div class="title">EJEMPLAR PARA EL PRESTADOR DEL SERVICIO</div>
    <table class="b hdr"><tr>
      <td style="width:50%;border-right:1.4px solid #000;padding:0"><table class="inner"><tr>
        <td class="c" style="width:60%">{logo_html}<div class="lbl">Nombre del taller</div><div class="val big">{e(empresa.get("nombre",""))}</div></td>
        {cell("CIF", empresa.get("nif",""))}</tr>
        <tr>{cell("Dirección", dir_taller)}{cell("RIIA", "")}</tr>
        <tr>{cell("Mail", empresa.get("email",""))}{cell("Teléfono", empresa.get("telefono",""))}</tr>
        <tr>{cell("", "")}{cell("Fax", "")}</tr></table></td>
      <td style="width:50%;padding:0;vertical-align:top">
        <div class="rgtitle">RESGUARDO DE DEPÓSITO SIN PRESUPUESTO Nº: <b>{e(orden.get("numero",""))}</b></div>
        <table class="inner">
          <tr>{cell("Titular del vehículo", cliente.get("nombre") or orden.get("cliente_nombre",""))}{cell("CIF/DNI titular", cliente.get("nif",""))}</tr>
          <tr>{cell("Persona solicitante","")}{cell("CIF/DNI solicitante","")}</tr>
          <tr>{cell("Dirección titular", cliente.get("direccion",""))}{cell("Teléfono", cliente.get("telefono",""))}</tr>
          <tr>{cell("Mail", cliente.get("email",""))}{cell("Fax","")}</tr>
        </table></td>
    </tr></table>

    <table class="b body"><tr>
      <td style="width:48%;border-right:1.4px solid #000;padding:0;vertical-align:top">
        <div class="sec">REPARACIONES A REALIZAR</div>
        <table class="rep"><tr class="hr"><th style="width:16px">Nº</th><th style="text-align:left">DESCRIPCIÓN</th><th style="width:52px">MANO DE OBRA</th><th style="width:48px">MATERIALES</th></tr>{filas}</table>
      </td>
      <td style="width:52%;padding:0;vertical-align:top">
        <table class="inner" style="border-bottom:1.4px solid #000"><tr>
          <td rowspan="4" style="width:150px;text-align:center;border-right:1px solid #000">{car}</td>
          {cell("Fecha", fmtf(orden.get("fecha_entrada")))}{cell("Matrícula", orden.get("vehiculo_matricula") or vehiculo.get("matricula",""))}{cell("Marca", vehiculo.get("marca",""))}</tr>
          <tr>{cell("Km", vehiculo.get("kilometros") if vehiculo.get("kilometros") is not None else "")}<td class="c"><div class="lbl">Seguro</div><div class="val">☐ SÍ &nbsp; ☐ NO</div></td>{cell("Modelo", vehiculo.get("modelo",""))}</tr>
          <tr><td class="c" colspan="3"><span class="lbl">Combustible R</span> <span class="val">{e(vehiculo.get("combustible",""))}</span></td></tr>
          <tr><td class="c" colspan="3"><div class="lbl">Observaciones</div><div class="val" style="font-weight:normal">{e(obs)}</div></td></tr>
        </table>
        <div class="renuncia">
          <div class="rtit">RENUNCIA A LA ELABORACIÓN<br>DE PRESUPUESTO PREVIO</div>
          <div class="rtxt">EL CLIENTE TIENE DERECHO A LA ELABORACIÓN DE UN PRESUPUESTO PREVIO. MEDIANTE LA PRESENTE FIRMA EL USUARIO RENUNCIA A LA ELABORACIÓN DE PRESUPUESTO PREVIO Y AUTORIZA A REALIZAR LOS TRABAJOS NECESARIOS PARA LA REPARACIÓN DEL VEHÍCULO Y/O SERVICIOS SOLICITADOS CONFORME A LO REFLEJADO EN ESTE RESGUARDO DE DEPÓSITO.</div>
          <table style="width:100%;margin-top:8px"><tr>
            <td style="width:50%"><b>EL PRESTADOR DEL SERVICIO</b><br>{e(empresa.get("nombre",""))}</td>
            <td style="width:50%;text-align:center"><b style="display:block;text-align:left">CONFORME CLIENTE</b>{firma_html}</td>
          </tr></table>
        </div>
        <div class="fentrega"><b>FECHA PREVISTA DE ENTREGA DEL VEHÍCULO REPARADO</b> <span style="float:right"><b>{e(fmtf(orden.get("fecha_entrega_estimada")))}</b></span></div>
        <div class="autoriz"><b>EL CLIENTE CON LA FIRMA ANTERIOR AUTORIZA AL TALLER A:</b>
          <table style="width:100%;margin-top:4px"><tr>
            <td style="width:50%">☐ REALIZAR DESPLAZAMIENTOS DE DIAGNÓSTICO.</td>
            <td style="width:50%">☐ UTILIZAR ELEMENTOS, EQUIPOS O CONJUNTOS USADOS O NO ESPECÍFICOS (ART. 9 Y 10 DECRETO 9/2003).</td></tr>
            <tr><td>☐ UTILIZAR ELEMENTOS, EQUIPOS O CONJUNTOS RECONSTRUIDOS (ART. 9 Y 10 DECRETO 9/2003).</td>
            <td>☐ RENUNCIA A RETIRAR ELEMENTOS SUSTITUIDOS TRAS REPARACIÓN.</td></tr></table>
        </div>
      </td>
    </tr></table>

    <table class="b legal"><tr>
      <td style="width:50%;border-right:1.4px solid #000"><b>Protección de Datos de Carácter Personal:</b> con la firma del presente usted presta su consentimiento para que sus datos sean tratados mientras que no comunique lo contrario por este taller, con la finalidad de gestión contable/administrativa de los servicios. Podrá ejercitar sus derechos de acceso, rectificación, supresión, oposición, y los demás reconocidos en esta norma, enviando solicitud a la dirección indicada, remitiendo copia de su DNI. Puede ejercitar el derecho a presentar una reclamación ante la Agencia Española de Protección de Datos.</td>
      <td style="width:50%">SI TRANSCURRIDOS TRES DÍAS DESDE LA PUESTA EN CONOCIMIENTO DEL CLIENTE DE LA FINALIZACIÓN DE LOS TRABAJOS DE ELABORACIÓN DEL PRESUPUESTO O REPARACIÓN DEL VEHÍCULO, NO PROCEDA EL CLIENTE AL PRONUNCIAMIENTO SOBRE LA ACEPTACIÓN O NO DEL PRESUPUESTO O A LA RETIRADA DEL VEHÍCULO, SE DEVENGARÁN UNOS GASTOS DIARIOS DE ESTANCIA DE __________ € MÁS IVA.</td>
    </tr></table>
    '''

    css = """
    @page { size: A4 portrait; margin: 8mm; }
    * { box-sizing:border-box; font-family: Arial, Helvetica, sans-serif; }
    body { margin:0; color:#000; font-size:9px; }
    table { border-collapse:collapse; width:100%; }
    .b { border:1.4px solid #000; }
    .title { text-align:center; font-size:18px; font-weight:bold; padding:5px 0; }
    .hdr td { vertical-align:top; }
    .inner { width:100%; }
    .inner td.c { border-right:.8px solid #000; border-bottom:.8px solid #000; padding:2px 6px; vertical-align:top; }
    .lbl { font-size:6.5px; text-transform:uppercase; color:#333; }
    .val { font-size:10px; font-weight:bold; min-height:12px; }
    .val.big { font-size:13px; }
    .rgtitle { border-bottom:.8px solid #000; padding:3px 6px; font-size:9px; font-weight:bold; }
    .body td { vertical-align:top; }
    .sec { text-align:center; font-weight:bold; font-size:10px; padding:3px 0; border-bottom:.8px solid #000; }
    .rep th { font-size:7px; border:.8px solid #000; background:#f0f0f0; padding:2px; }
    .rep td { border-right:.8px solid #000; border-bottom:1px dotted #aaa; font-size:8.5px; padding:1px 4px; }
    .rep td.n { text-align:center; color:#555; }
    .rep td.m { text-align:right; }
    .renuncia { border:1.2px solid #000; margin:6px; padding:8px 10px; }
    .rtit { text-align:center; font-size:15px; font-weight:bold; line-height:1.1; margin-bottom:6px; }
    .rtxt { font-size:8px; font-weight:bold; text-align:justify; line-height:1.4; }
    .fentrega { border-top:1.2px solid #000; border-bottom:1.2px solid #000; padding:4px 8px; font-size:8.5px; }
    .autoriz { padding:5px 8px; font-size:8px; }
    .legal td { padding:5px 8px; font-size:7px; text-align:justify; line-height:1.4; }
    """
    return f"<!doctype html><html><head><meta charset='utf-8'><style>{css}</style></head><body>{body}</body></html>"


@api_router.get("/taller/ordenes/{oid}/hoja-entrada.pdf")
async def hoja_entrada_pdf(oid: str):
    orden = await db.ordenes_trabajo.find_one({"id": oid}, {"_id": 0})
    if not orden:
        raise HTTPException(404, "Orden no encontrada")
    vehiculo = await db.vehiculos.find_one({"id": orden.get("vehiculo_id")}, {"_id": 0}) or {}
    cliente = await db.contactos.find_one({"id": orden.get("cliente_id") or vehiculo.get("cliente_id")}, {"_id": 0}) or {}
    cfg = await _get_ajustes()
    empresa = cfg.get("empresa", {}) or {}
    html = _build_resguardo_html(orden, vehiculo, cliente, empresa)
    from weasyprint import HTML as _WHTML
    pdf = await asyncio.to_thread(lambda: _WHTML(string=html).write_pdf())
    return Response(content=pdf, media_type="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="hoja-entrada-{orden.get("numero","")}.pdf"'})


@api_router.get("/taller/ordenes/{oid}/hoja-entrada.html")
async def hoja_entrada_html(oid: str):
    orden = await db.ordenes_trabajo.find_one({"id": oid}, {"_id": 0})
    if not orden:
        raise HTTPException(404, "Orden no encontrada")
    vehiculo = await db.vehiculos.find_one({"id": orden.get("vehiculo_id")}, {"_id": 0}) or {}
    cliente = await db.contactos.find_one({"id": orden.get("cliente_id") or vehiculo.get("cliente_id")}, {"_id": 0}) or {}
    cfg = await _get_ajustes()
    empresa = cfg.get("empresa", {}) or {}
    html = _build_resguardo_html(orden, vehiculo, cliente, empresa)
    return Response(content=html, media_type="text/html; charset=utf-8")


def _build_parte_html(orden, vehiculo, cliente, empresa):
    def e(v):
        return str(v if v is not None else "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def money(n):
        try:
            v = float(n or 0)
        except Exception:
            v = 0.0
        return f"{v:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")

    def fmtf(f):
        if not f:
            return datetime.now(timezone.utc).strftime("%d/%m/%Y")
        try:
            return datetime.fromisoformat(str(f)[:19]).strftime("%d/%m/%Y")
        except Exception:
            return str(f)

    ACCENT = "#4338ca"
    logo = _img_data_uri(empresa.get("logo", ""))
    dir_taller = " · ".join([x for x in [empresa.get("direccion", ""),
                             " ".join([x for x in [empresa.get("codigo_postal", ""), empresa.get("ciudad", "")] if x])] if x])
    trabajos = ", ".join([_TIPOS_TRABAJO.get(t, t) for t in (orden.get("tipos_trabajo") or [])])

    filas = ""
    for l in (orden.get("lineas") or []):
        cant = float(l.get("cantidad", 0) or 0)
        pu = float(l.get("precio_unitario", 0) or 0)
        base = cant * pu * (1 - float(l.get("descuento", 0) or 0) / 100)
        tot = l.get("total")
        if tot is None:
            tot = base * (1 + float(l.get("tipo_iva", 0) or 0) / 100)
        filas += (f'<tr><td class="ld">{e(l.get("descripcion",""))}</td>'
                  f'<td class="lr">{e(l.get("cantidad",""))} {e(l.get("unidad","ud"))}</td>'
                  f'<td class="lr">{money(pu)}</td>'
                  f'<td class="lr" style="font-weight:600">{money(tot)}</td></tr>')
    if not filas:
        filas = '<tr><td colspan="4" style="padding:16px;text-align:center;color:#a1a1aa">Sin líneas</td></tr>'

    fotos = ""
    imgs = [f for f in (orden.get("fotos") or []) if "pdf" not in (f.get("content_type") or "")]
    if imgs:
        cells = ""
        for f in imgs[:12]:
            uri = _img_data_uri(f.get("path", ""))
            if uri:
                cells += f'<div class="fcell"><img src="{uri}"/></div>'
        if cells:
            fotos = (f'<div class="fttl">Reportaje fotográfico ({len(imgs)})</div>'
                     f'<div class="fgrid">{cells}</div>')

    logo_html = (f'<img src="{logo}" style="max-height:56px;max-width:190px;object-fit:contain"/>'
                 if logo else f'<div style="font-size:20px;font-weight:800">{e(empresa.get("nombre","Taller"))}</div>')

    css = """
    @page { size: A4 portrait; margin: 14mm; }
    * { box-sizing:border-box; font-family:'Helvetica','Arial',sans-serif; }
    body { margin:0; color:#18181b; font-size:12px; }
    .head { display:flex; justify-content:space-between; align-items:flex-start;
            border-bottom:3px solid __AC__; padding-bottom:14px; margin-bottom:20px; }
    .hinfo { font-size:10px; color:#52525b; line-height:1.5; margin-top:4px; }
    .doctit { font-size:24px; font-weight:800; text-align:right; }
    .docref { font-family:monospace; font-size:12px; color:#3f3f46; text-align:right; margin-top:3px; }
    .docdate { font-size:11px; color:#71717a; text-align:right; }
    .sec { font-size:11px; font-weight:700; color:__AC__; text-transform:uppercase;
           letter-spacing:.06em; margin:0 0 8px; }
    .grid { display:grid; grid-template-columns:repeat(3,1fr); gap:12px;
            background:#fafafa; border:1px solid #eee; border-radius:8px; padding:14px; margin-bottom:18px; }
    .g-lbl { font-size:9px; text-transform:uppercase; letter-spacing:.08em; color:#a1a1aa; }
    .g-val { font-weight:600; font-size:12px; }
    .desc { font-size:12px; color:#3f3f46; margin-bottom:16px; }
    table.lin { width:100%; border-collapse:collapse; }
    table.lin th { background:#f4f4f5; padding:8px; font-size:9px; text-transform:uppercase;
                   color:#71717a; text-align:right; }
    table.lin th:first-child { text-align:left; }
    table.lin td { padding:7px 8px; border-bottom:1px solid #f1f1f4; }
    td.ld { text-align:left; } td.lr { text-align:right; }
    .tot { width:250px; margin-left:auto; margin-top:14px; }
    .tot tr td { padding:4px 8px; }
    .tot .k { color:#71717a; } .tot .v { text-align:right; }
    .tot .grand td { border-top:2px solid __AC__; font-weight:800; font-size:14px; }
    .tot .grand .v { color:__AC__; }
    .fttl { font-size:11px; font-weight:700; color:__AC__; text-transform:uppercase;
            letter-spacing:.06em; margin:22px 0 10px; }
    .fgrid { display:grid; grid-template-columns:repeat(3,1fr); gap:8px; }
    .fcell { border:1px solid #e4e4e7; border-radius:6px; overflow:hidden; height:130px; }
    .fcell img { width:100%; height:100%; object-fit:cover; }
    .firmas { display:flex; justify-content:space-between; gap:40px; margin-top:44px; }
    .firmas div { flex:1; border-top:1px solid #d4d4d8; padding-top:6px; font-size:11px; color:#71717a; }
    """.replace("__AC__", ACCENT)

    def g(lbl, val):
        if not val and val != 0:
            return ""
        return f'<div><div class="g-lbl">{e(lbl)}</div><div class="g-val">{e(val)}</div></div>'

    km = vehiculo.get("kilometros")
    body = f'''
    <div class="head">
      <div>{logo_html}<div class="hinfo">
        {('<div style="font-weight:600;color:#18181b">'+e(empresa.get("nombre",""))+'</div>') if empresa.get("nombre") else ''}
        {('<div>NIF: '+e(empresa.get("nif",""))+'</div>') if empresa.get("nif") else ''}
        {('<div>'+e(dir_taller)+'</div>') if dir_taller else ''}
        {('<div>Tel: '+e(empresa.get("telefono",""))+'</div>') if empresa.get("telefono") else ''}
      </div></div>
      <div><div class="doctit">Parte de trabajo</div>
        <div class="docref">{e(orden.get("numero","—"))}</div>
        <div class="docdate">{fmtf(orden.get("fecha_entrada"))}</div></div>
    </div>
    <div class="sec">Vehículo</div>
    <div class="grid">
      {g("Matrícula", orden.get("vehiculo_matricula") or vehiculo.get("matricula",""))}
      {g("Marca / Modelo", " ".join([x for x in [vehiculo.get("marca",""), vehiculo.get("modelo","")] if x]))}
      {g("Kilómetros", (str(km)+" km") if km is not None else "")}
      {g("Cliente", cliente.get("nombre") or orden.get("cliente_nombre",""))}
      {g("Tipo de trabajo", trabajos)}
      {g("Estado", orden.get("estado",""))}
    </div>
    {('<div class="sec">Descripción</div><div class="desc">'+e(orden.get("descripcion",""))+'</div>') if orden.get("descripcion") else ''}
    <table class="lin"><thead><tr>
      <th>Concepto</th><th>Cant.</th><th>Precio</th><th>Total</th>
    </tr></thead><tbody>{filas}</tbody></table>
    <table class="tot">
      <tr><td class="k">Base</td><td class="v">{money(orden.get("base"))}</td></tr>
      <tr><td class="k">IVA</td><td class="v">{money(orden.get("cuota_iva"))}</td></tr>
      <tr class="grand"><td>TOTAL</td><td class="v">{money(orden.get("total"))}</td></tr>
    </table>
    {fotos}
    <div class="firmas"><div>Firma del taller</div><div>Firma del cliente (conforme)</div></div>
    '''
    return f"<!doctype html><html><head><meta charset='utf-8'><style>{css}</style></head><body>{body}</body></html>"


@api_router.get("/taller/ordenes/{oid}/parte-trabajo.pdf")
async def parte_trabajo_pdf(oid: str):
    orden = await db.ordenes_trabajo.find_one({"id": oid}, {"_id": 0})
    if not orden:
        raise HTTPException(404, "Orden no encontrada")
    vehiculo = await db.vehiculos.find_one({"id": orden.get("vehiculo_id")}, {"_id": 0}) or {}
    cliente = await db.contactos.find_one({"id": orden.get("cliente_id") or vehiculo.get("cliente_id")}, {"_id": 0}) or {}
    cfg = await _get_ajustes()
    empresa = cfg.get("empresa", {}) or {}
    html = _build_parte_html(orden, vehiculo, cliente, empresa)
    from weasyprint import HTML as _WHTML
    pdf = await asyncio.to_thread(lambda: _WHTML(string=html).write_pdf())
    return Response(content=pdf, media_type="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="parte-{orden.get("numero","")}.pdf"'})


@api_router.get("/taller/ordenes/{oid}/parte-trabajo.html")
async def parte_trabajo_html(oid: str):
    orden = await db.ordenes_trabajo.find_one({"id": oid}, {"_id": 0})
    if not orden:
        raise HTTPException(404, "Orden no encontrada")
    vehiculo = await db.vehiculos.find_one({"id": orden.get("vehiculo_id")}, {"_id": 0}) or {}
    cliente = await db.contactos.find_one({"id": orden.get("cliente_id") or vehiculo.get("cliente_id")}, {"_id": 0}) or {}
    cfg = await _get_ajustes()
    empresa = cfg.get("empresa", {}) or {}
    html = _build_parte_html(orden, vehiculo, cliente, empresa)
    return Response(content=html, media_type="text/html; charset=utf-8")




# ---- Sesiones de subida por QR (públicas) ----
class FotoSesionInput(BaseModel):
    tipo: str
    entidad_id: str


@api_router.post("/taller/foto-sesion")
async def crear_foto_sesion(data: FotoSesionInput):
    if data.tipo not in _TALLER_COLS:
        raise HTTPException(400, "Tipo no válido")
    token = uuid.uuid4().hex[:10]
    await db.foto_sesiones.insert_one({
        "token": token, "tipo": data.tipo, "entidad_id": data.entidad_id, "created_at": now_iso(),
    })
    return {"token": token}


@api_router.get("/taller/subida/{token}")
async def info_subida(token: str):
    s = await db.foto_sesiones.find_one({"token": token}, {"_id": 0})
    if not s:
        raise HTTPException(404, "Enlace de subida no válido o caducado")
    col = _TALLER_COLS[s["tipo"]]
    ent = await db[col].find_one({"id": s["entidad_id"]}, {"_id": 0})
    label = ""
    if ent:
        label = ent.get("matricula") or ent.get("vehiculo_matricula") or ent.get("numero") or ""
    total = len(ent.get("fotos", [])) if ent else 0
    return {"tipo": s["tipo"], "label": label, "total": total}


@api_router.post("/taller/subida/{token}")
async def subir_foto_token(token: str, file: UploadFile = File(...)):
    s = await db.foto_sesiones.find_one({"token": token})
    if not s:
        raise HTTPException(404, "Enlace de subida no válido o caducado")
    col = _TALLER_COLS[s["tipo"]]
    foto = await _append_foto(col, s["entidad_id"], file)
    ent = await db[col].find_one({"id": s["entidad_id"]}, {"_id": 0})
    return {"ok": True, "total": len(ent.get("fotos", [])) if ent else 0, "foto": foto}


# ---- Compañías de seguros ----
class CompaniaInput(BaseModel):
    nombre: str
    cif: str = ""
    telefono: str = ""
    email: str = ""


class Compania(CompaniaInput):
    id: str = Field(default_factory=new_id)
    created_at: str = Field(default_factory=now_iso)


@api_router.post("/taller/companias")
async def crear_compania(data: CompaniaInput):
    c = Compania(**data.model_dump())
    await db.companias.insert_one(c.model_dump())
    return c.model_dump()


@api_router.get("/taller/companias")
async def listar_companias():
    return await db.companias.find({}, {"_id": 0}).sort("nombre", 1).to_list(1000)


@api_router.put("/taller/companias/{cid}")
async def actualizar_compania(cid: str, data: CompaniaInput):
    res = await db.companias.update_one({"id": cid}, {"$set": data.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(404, "Compañía no encontrada")
    return await db.companias.find_one({"id": cid}, {"_id": 0})


@api_router.delete("/taller/companias/{cid}")
async def eliminar_compania(cid: str):
    await db.companias.delete_one({"id": cid})
    return {"ok": True}


# ---- Peritajes ----
ESTADOS_PERITAJE = ['pendiente', 'valorado', 'aprobado', 'rechazado']


class DanioItem(BaseModel):
    descripcion: str = ""
    importe: float = 0


class PeritajeInput(BaseModel):
    vehiculo_id: str = ""
    vehiculo_matricula: str = ""
    cliente_id: str = ""
    cliente_nombre: str = ""
    compania: str = ""
    poliza: str = ""
    siniestro: str = ""
    descripcion: str = ""
    danios: List[DanioItem] = []
    estado: Literal['pendiente', 'valorado', 'aprobado', 'rechazado'] = 'pendiente'
    fecha: str = ""


class Peritaje(PeritajeInput):
    id: str = Field(default_factory=new_id)
    numero: str = ""
    importe_total: float = 0
    fotos: List[dict] = []
    created_at: str = Field(default_factory=now_iso)


async def _rellena_peritaje(d: dict) -> dict:
    if d.get("vehiculo_id"):
        v = await db.vehiculos.find_one({"id": d["vehiculo_id"]}, {"_id": 0})
        if v:
            d["vehiculo_matricula"] = v.get("matricula", "")
            if not d.get("cliente_id"):
                d["cliente_id"] = v.get("cliente_id", "")
                d["cliente_nombre"] = v.get("cliente_nombre", "")
    if d.get("cliente_id") and not d.get("cliente_nombre"):
        cli = await db.contactos.find_one({"id": d["cliente_id"]}, {"_id": 0})
        if cli:
            d["cliente_nombre"] = cli.get("nombre", "")
    d["importe_total"] = round(sum(float(x.get("importe", 0) or 0) for x in d.get("danios", [])), 2)
    return d


@api_router.post("/taller/peritajes")
async def crear_peritaje(data: PeritajeInput):
    d = await _rellena_peritaje(data.model_dump())
    n = await _next_seq("peritaje")
    d["numero"] = f"PER-{n:06d}"
    p = Peritaje(**d)
    await db.peritajes.insert_one(p.model_dump())
    return p.model_dump()


@api_router.get("/taller/peritajes")
async def listar_peritajes(vehiculo_id: Optional[str] = None):
    q = {}
    if vehiculo_id:
        q["vehiculo_id"] = vehiculo_id
    return await db.peritajes.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)


@api_router.get("/taller/peritajes/{pid}")
async def obtener_peritaje(pid: str):
    doc = await db.peritajes.find_one({"id": pid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Peritaje no encontrado")
    return doc


@api_router.put("/taller/peritajes/{pid}")
async def actualizar_peritaje(pid: str, data: PeritajeInput):
    existing = await db.peritajes.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Peritaje no encontrado")
    d = await _rellena_peritaje(data.model_dump())
    d["numero"] = existing.get("numero", "")
    await db.peritajes.update_one({"id": pid}, {"$set": d})
    return await db.peritajes.find_one({"id": pid}, {"_id": 0})


@api_router.patch("/taller/peritajes/{pid}/estado")
async def estado_peritaje(pid: str, estado: str = Form(...)):
    if estado not in ESTADOS_PERITAJE:
        raise HTTPException(400, "Estado inválido")
    res = await db.peritajes.update_one({"id": pid}, {"$set": {"estado": estado}})
    if res.matched_count == 0:
        raise HTTPException(404, "Peritaje no encontrado")
    return await db.peritajes.find_one({"id": pid}, {"_id": 0})


@api_router.delete("/taller/peritajes/{pid}")
async def eliminar_peritaje(pid: str):
    await db.peritajes.delete_one({"id": pid})
    return {"ok": True}


# ---------------------------------------------------------------------------
# TALLER — FASE 3: Citas + Vehículos de cortesía (préstamos)
# ---------------------------------------------------------------------------
ESTADOS_CITA = ['pendiente', 'confirmada', 'realizada', 'cancelada']


class CitaInput(BaseModel):
    vehiculo_id: str = ""
    vehiculo_matricula: str = ""
    cliente_id: str = ""
    cliente_nombre: str = ""
    fecha: str = ""          # ISO datetime-local (YYYY-MM-DDTHH:MM)
    duracion_min: int = 60
    motivo: str = ""
    tipo_trabajo: str = ""   # chapa / pintura / mecanica / revision
    estado: Literal['pendiente', 'confirmada', 'realizada', 'cancelada'] = 'pendiente'
    notas: str = ""


class Cita(CitaInput):
    id: str = Field(default_factory=new_id)
    token: str = Field(default_factory=lambda: secrets.token_urlsafe(9))
    created_at: str = Field(default_factory=now_iso)


async def _rellena_cita(d: dict) -> dict:
    if d.get("vehiculo_id"):
        v = await db.vehiculos.find_one({"id": d["vehiculo_id"]}, {"_id": 0})
        if v:
            d["vehiculo_matricula"] = v.get("matricula", "")
            if not d.get("cliente_id"):
                d["cliente_id"] = v.get("cliente_id", "")
                d["cliente_nombre"] = v.get("cliente_nombre", "")
    if d.get("cliente_id") and not d.get("cliente_nombre"):
        cli = await db.contactos.find_one({"id": d["cliente_id"]}, {"_id": 0})
        if cli:
            d["cliente_nombre"] = cli.get("nombre", "")
    return d


@api_router.post("/taller/citas")
async def crear_cita(data: CitaInput):
    d = await _rellena_cita(data.model_dump())
    c = Cita(**d)
    await db.citas.insert_one(c.model_dump())
    return c.model_dump()


@api_router.get("/taller/citas")
async def listar_citas(vehiculo_id: Optional[str] = None, desde: Optional[str] = None, hasta: Optional[str] = None):
    q = {}
    if vehiculo_id:
        q["vehiculo_id"] = vehiculo_id
    if desde or hasta:
        rng = {}
        if desde:
            rng["$gte"] = desde
        if hasta:
            rng["$lte"] = hasta
        q["fecha"] = rng
    return await db.citas.find(q, {"_id": 0}).sort("fecha", 1).to_list(3000)


@api_router.put("/taller/citas/{cid}")
async def actualizar_cita(cid: str, data: CitaInput):
    d = await _rellena_cita(data.model_dump())
    res = await db.citas.update_one({"id": cid}, {"$set": d})
    if res.matched_count == 0:
        raise HTTPException(404, "Cita no encontrada")
    return await db.citas.find_one({"id": cid}, {"_id": 0})


@api_router.patch("/taller/citas/{cid}/estado")
async def estado_cita(cid: str, estado: str = Form(...)):
    if estado not in ESTADOS_CITA:
        raise HTTPException(400, "Estado inválido")
    res = await db.citas.update_one({"id": cid}, {"$set": {"estado": estado}})
    if res.matched_count == 0:
        raise HTTPException(404, "Cita no encontrada")
    return await db.citas.find_one({"id": cid}, {"_id": 0})


@api_router.delete("/taller/citas/{cid}")
async def eliminar_cita(cid: str):
    await db.citas.delete_one({"id": cid})
    return {"ok": True}


# ---------------------------------------------------------------------------
# NOTIFICACIONES — Recordatorios de citas (Email vía Resend / WhatsApp vía Twilio)
# ---------------------------------------------------------------------------
def _fmt_tel_e164(tel: str) -> str:
    t = re.sub(r"[^\d+]", "", tel or "")
    if not t:
        return ""
    if not t.startswith("+"):
        t = "+34" + t.lstrip("0")
    return t


def _fecha_hora_cita(fecha: str):
    if not fecha:
        return "", ""
    try:
        dt = datetime.fromisoformat(fecha)
        return dt.strftime("%d/%m/%Y"), dt.strftime("%H:%M")
    except Exception:
        parts = fecha.split("T")
        f = parts[0]
        try:
            y, m, d = f.split("-")
            f = f"{d}/{m}/{y}"
        except Exception:
            pass
        return f, (parts[1][:5] if len(parts) > 1 else "")


def _rellena_plantilla(txt: str, ctx: dict) -> str:
    out = txt or ""
    for k, v in ctx.items():
        out = out.replace("{" + k + "}", str(v or ""))
    return out


async def _enviar_email(email_cfg: dict, to_email: str, asunto: str, cuerpo: str):
    resend.api_key = email_cfg.get("api_key") or ""
    frm = email_cfg.get("from_email") or ""
    if email_cfg.get("from_nombre"):
        frm = f'{email_cfg["from_nombre"]} <{email_cfg["from_email"]}>'
    html = ("<div style=\"font-family:Arial,Helvetica,sans-serif;font-size:14px;color:#18181b;"
            "line-height:1.6\">" + (cuerpo or "").replace("\n", "<br>") + "</div>")
    params = {"from": frm, "to": [to_email], "subject": asunto, "html": html}
    return await asyncio.to_thread(resend.Emails.send, params)


async def _enviar_whatsapp(wa_cfg: dict, to_number: str, body: str):
    cli = TwilioClient(wa_cfg.get("account_sid") or "", wa_cfg.get("auth_token") or "")
    frm = wa_cfg.get("from_number") or ""

    def _send():
        return cli.messages.create(
            from_=f"whatsapp:{frm}" if not frm.startswith("whatsapp:") else frm,
            to=f"whatsapp:{to_number}",
            body=body,
        )
    msg = await asyncio.to_thread(_send)
    return msg.sid


async def _enviar_recordatorio_cita(cita: dict, canal: Optional[str] = None, marcar_auto: bool = False):
    cfg = await _get_ajustes()
    notif = _merge_notif(cfg)
    empresa = cfg.get("empresa", {})
    rec = notif.get("recordatorios", {})
    email_cfg, wa_cfg = notif.get("email", {}), notif.get("whatsapp", {})

    cli = None
    if cita.get("cliente_id"):
        cli = await db.contactos.find_one({"id": cita["cliente_id"]}, {"_id": 0})
    email_dest = ((cli or {}).get("email") or "").strip()
    tel_dest = _fmt_tel_e164((cli or {}).get("telefono") or "")

    f, h = _fecha_hora_cita(cita.get("fecha", ""))
    token = cita.get("token")
    if not token:
        token = secrets.token_urlsafe(9)
        await db.citas.update_one({"id": cita["id"]}, {"$set": {"token": token}})
    app_url = cfg.get("app_url") or ""
    enlace = f"{app_url}/cita/{token}" if app_url else ""
    ctx = {
        "cliente": cita.get("cliente_nombre") or (cli or {}).get("nombre") or "cliente",
        "empresa": empresa.get("nombre") or "nuestro taller",
        "fecha": f, "hora": h,
        "matricula": cita.get("vehiculo_matricula") or "—",
        "motivo": cita.get("motivo") or "",
        "telefono": empresa.get("telefono") or "",
        "enlace": enlace,
    }

    canal = canal or rec.get("canal") or "email"
    do_email = canal in ("email", "ambos")
    do_wa = canal in ("whatsapp", "ambos")
    resultados, enviado = {}, False

    if do_email:
        if not email_cfg.get("activo") or not email_cfg.get("api_key") or not email_cfg.get("from_email"):
            resultados["email"] = {"ok": False, "error": "Email no configurado en Ajustes"}
        elif not email_dest:
            resultados["email"] = {"ok": False, "error": "El cliente no tiene email"}
        else:
            try:
                asunto = _rellena_plantilla(rec.get("email_asunto"), ctx)
                cuerpo_tpl = rec.get("email_cuerpo") or ""
                cuerpo = _rellena_plantilla(cuerpo_tpl, ctx)
                if enlace and "{enlace}" not in cuerpo_tpl:
                    cuerpo += f"\n\nConfirmar o cancelar su cita: {enlace}"
                r = await _enviar_email(email_cfg, email_dest, asunto, cuerpo)
                resultados["email"] = {"ok": True, "id": (r or {}).get("id"), "destino": email_dest}
                enviado = True
            except Exception as e:
                resultados["email"] = {"ok": False, "error": str(e)}

    if do_wa:
        if not wa_cfg.get("activo") or not wa_cfg.get("account_sid") or not wa_cfg.get("auth_token") or not wa_cfg.get("from_number"):
            resultados["whatsapp"] = {"ok": False, "error": "WhatsApp no configurado en Ajustes"}
        elif not tel_dest:
            resultados["whatsapp"] = {"ok": False, "error": "El cliente no tiene teléfono"}
        else:
            try:
                texto_tpl = rec.get("whatsapp_texto") or ""
                texto = _rellena_plantilla(texto_tpl, ctx)
                if enlace and "{enlace}" not in texto_tpl:
                    texto += f"\nConfirmar o cancelar: {enlace}"
                sid = await _enviar_whatsapp(wa_cfg, tel_dest, texto)
                resultados["whatsapp"] = {"ok": True, "sid": sid, "destino": tel_dest}
                enviado = True
            except Exception as e:
                resultados["whatsapp"] = {"ok": False, "error": str(e)}

    upd = {"recordatorio_resultado": resultados}
    if enviado:
        upd["recordatorio_enviado_at"] = now_iso()
        upd["recordatorio_canales"] = [k for k, v in resultados.items() if v.get("ok")]
    if marcar_auto:
        upd["recordatorio_auto_at"] = now_iso()
    await db.citas.update_one({"id": cita["id"]}, {"$set": upd})
    return {"enviado": enviado, "resultados": resultados}


@api_router.post("/taller/citas/{cid}/recordatorio")
async def enviar_recordatorio(cid: str, canal: Optional[str] = Form(None), base_url: Optional[str] = Form(None)):
    cita = await db.citas.find_one({"id": cid}, {"_id": 0})
    if not cita:
        raise HTTPException(404, "Cita no encontrada")
    if base_url:
        await db.ajustes.update_one({"_id": "config"}, {"$set": {"app_url": base_url.rstrip("/")}}, upsert=True)
    res = await _enviar_recordatorio_cita(cita, canal=canal)
    return res


# --- Confirmación pública de citas (enlace del recordatorio, sin login) ---
@api_router.get("/public/cita/{token}")
async def cita_publica(token: str):
    cita = await db.citas.find_one({"token": token}, {"_id": 0})
    if not cita:
        raise HTTPException(404, "Cita no encontrada")
    cfg = await _get_ajustes()
    f, h = _fecha_hora_cita(cita.get("fecha", ""))
    return {
        "fecha": f, "hora": h,
        "vehiculo_matricula": cita.get("vehiculo_matricula") or "",
        "cliente_nombre": cita.get("cliente_nombre") or "",
        "motivo": cita.get("motivo") or "",
        "estado": cita.get("estado") or "pendiente",
        "empresa_nombre": (cfg.get("empresa", {}) or {}).get("nombre") or "el taller",
        "empresa_telefono": (cfg.get("empresa", {}) or {}).get("telefono") or "",
    }


@api_router.post("/public/cita/{token}/responder")
async def responder_cita(token: str, accion: str = Form(...)):
    if accion not in ("confirmar", "cancelar"):
        raise HTTPException(400, "Acción inválida")
    nuevo = "confirmada" if accion == "confirmar" else "cancelada"
    res = await db.citas.update_one(
        {"token": token},
        {"$set": {"estado": nuevo, "respuesta_cliente_at": now_iso(), "respuesta_cliente": nuevo}},
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Cita no encontrada")
    return {"ok": True, "estado": nuevo}


class TestNotifInput(BaseModel):
    canal: Literal['email', 'whatsapp']
    destino: str


@api_router.post("/notificaciones/test")
async def probar_notificacion(data: TestNotifInput):
    cfg = await _get_ajustes()
    notif = _merge_notif(cfg)
    empresa = cfg.get("empresa", {})
    nombre = empresa.get("nombre") or "tu taller"
    if data.canal == "email":
        ec = notif.get("email", {})
        if not ec.get("api_key") or not ec.get("from_email"):
            raise HTTPException(400, "Configura la API key y el email remitente antes de probar")
        try:
            r = await _enviar_email(ec, data.destino.strip(),
                                    f"Prueba de recordatorios · {nombre}",
                                    f"Hola,\n\nEste es un email de prueba de {nombre}. La configuración de correo funciona correctamente.")
            return {"ok": True, "id": (r or {}).get("id")}
        except Exception as e:
            raise HTTPException(400, f"Error al enviar el email: {e}")
    else:
        wc = notif.get("whatsapp", {})
        if not wc.get("account_sid") or not wc.get("auth_token") or not wc.get("from_number"):
            raise HTTPException(400, "Configura SID, token y número antes de probar")
        try:
            sid = await _enviar_whatsapp(wc, _fmt_tel_e164(data.destino),
                                         f"Prueba de recordatorios de {nombre}. La configuración de WhatsApp funciona correctamente.")
            return {"ok": True, "sid": sid}
        except Exception as e:
            raise HTTPException(400, f"Error al enviar el WhatsApp: {e}")


async def _job_recordatorios():
    try:
        cfg = await _get_ajustes()
        rec = _merge_notif(cfg).get("recordatorios", {})
        if not rec.get("activo"):
            return
        horas = int(rec.get("horas_antes") or 24)
        now = datetime.now()
        ahora_iso = now.strftime("%Y-%m-%dT%H:%M")
        limite_iso = (now + timedelta(hours=horas)).strftime("%Y-%m-%dT%H:%M")
        q = {
            "fecha": {"$gte": ahora_iso, "$lte": limite_iso},
            "estado": {"$in": ["pendiente", "confirmada"]},
            "$or": [{"recordatorio_auto_at": {"$exists": False}},
                    {"recordatorio_auto_at": None}, {"recordatorio_auto_at": ""}],
        }
        citas = await db.citas.find(q, {"_id": 0}).to_list(500)
        for c in citas:
            await _enviar_recordatorio_cita(c, marcar_auto=True)
        if citas:
            logger.info(f"Recordatorios automáticos procesados: {len(citas)}")
    except Exception as e:
        logger.warning(f"Job recordatorios: {e}")


_scheduler = AsyncIOScheduler()


# ---- Vehículos de cortesía: préstamos ----
class PrestamoInput(BaseModel):
    vehiculo_id: str = ""            # vehículo de cortesía (tipo=cortesia)
    vehiculo_matricula: str = ""
    cliente_id: str = ""
    cliente_nombre: str = ""
    vehiculo_cliente_id: str = ""    # vehículo del cliente que está en reparación (opcional)
    fecha_entrega: str = ""
    fecha_devolucion_prevista: str = ""
    fecha_devolucion_real: str = ""
    km_entrega: Optional[int] = None
    km_devolucion: Optional[int] = None
    estado: Literal['activo', 'devuelto'] = 'activo'
    notas: str = ""


class Prestamo(PrestamoInput):
    id: str = Field(default_factory=new_id)
    contrato_path: str = ""
    contrato_filename: str = ""
    created_at: str = Field(default_factory=now_iso)


@api_router.post("/taller/prestamos")
async def crear_prestamo(data: PrestamoInput):
    d = data.model_dump()
    if d.get("vehiculo_id"):
        v = await db.vehiculos.find_one({"id": d["vehiculo_id"]}, {"_id": 0})
        if v:
            d["vehiculo_matricula"] = v.get("matricula", "")
    if d.get("cliente_id") and not d.get("cliente_nombre"):
        cli = await db.contactos.find_one({"id": d["cliente_id"]}, {"_id": 0})
        if cli:
            d["cliente_nombre"] = cli.get("nombre", "")
    p = Prestamo(**d)
    await db.prestamos.insert_one(p.model_dump())
    return p.model_dump()


@api_router.get("/taller/prestamos")
async def listar_prestamos(estado: Optional[str] = None):
    q = {}
    if estado:
        q["estado"] = estado
    return await db.prestamos.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)


@api_router.put("/taller/prestamos/{pid}")
async def actualizar_prestamo(pid: str, data: PrestamoInput):
    existing = await db.prestamos.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Préstamo no encontrado")
    d = data.model_dump()
    if d.get("vehiculo_id"):
        v = await db.vehiculos.find_one({"id": d["vehiculo_id"]}, {"_id": 0})
        if v:
            d["vehiculo_matricula"] = v.get("matricula", "")
    if d.get("cliente_id") and not d.get("cliente_nombre"):
        cli = await db.contactos.find_one({"id": d["cliente_id"]}, {"_id": 0})
        if cli:
            d["cliente_nombre"] = cli.get("nombre", "")
    await db.prestamos.update_one({"id": pid}, {"$set": d})
    return await db.prestamos.find_one({"id": pid}, {"_id": 0})


@api_router.post("/taller/prestamos/{pid}/contrato")
async def subir_contrato(pid: str, file: UploadFile = File(...)):
    contenido = await file.read()
    ct = file.content_type or ""
    if not (ct.startswith("image/") or ct == "application/pdf"):
        raise HTTPException(400, "El contrato debe ser una imagen o PDF")
    if len(contenido) > 15 * 1024 * 1024:
        raise HTTPException(400, "El archivo no debe superar 15 MB")
    media = await _guardar_media(contenido, file.filename, ct)
    res = await db.prestamos.update_one({"id": pid}, {"$set": {
        "contrato_path": media["path"], "contrato_filename": media["filename"]}})
    if res.matched_count == 0:
        raise HTTPException(404, "Préstamo no encontrado")
    return {"contrato_path": media["path"], "contrato_filename": media["filename"]}


@api_router.delete("/taller/prestamos/{pid}")
async def eliminar_prestamo(pid: str):
    await db.prestamos.delete_one({"id": pid})
    return {"ok": True}


@api_router.get("/taller/resumen")
async def taller_resumen():
    hoy = date.today().isoformat()
    ordenes = await db.ordenes_trabajo.find({}, {"_id": 0}).to_list(5000)
    por_estado = {e: 0 for e in ESTADOS_OT}
    for o in ordenes:
        est = o.get("estado", "recepcion")
        por_estado[est] = por_estado.get(est, 0) + 1
    abiertas = len([o for o in ordenes if o.get("estado") in ("recepcion", "en_curso")])
    citas_hoy = await db.citas.find(
        {"fecha": {"$gte": hoy + "T00:00", "$lte": hoy + "T23:59"}}, {"_id": 0}
    ).sort("fecha", 1).to_list(200)
    proximas = await db.citas.find(
        {"fecha": {"$gt": hoy + "T23:59"}}, {"_id": 0}
    ).sort("fecha", 1).to_list(6)
    peritajes_pend = await db.peritajes.count_documents({"estado": {"$in": ["pendiente", "valorado"]}})
    cortesias = await db.prestamos.find({"estado": "activo"}, {"_id": 0}).sort("created_at", -1).to_list(200)
    total_vehiculos = await db.vehiculos.count_documents({})
    ultimas = sorted(ordenes, key=lambda x: x.get("created_at", ""), reverse=True)[:6]
    return {
        "total_vehiculos": total_vehiculos,
        "ordenes_por_estado": por_estado,
        "ordenes_abiertas": abiertas,
        "citas_hoy": citas_hoy,
        "proximas_citas": proximas,
        "peritajes_pendientes": peritajes_pend,
        "cortesias_activas": cortesias,
        "ultimas_ordenes": ultimas,
    }










@api_router.get("/")
async def root():
    return {"message": "ERP Base API", "status": "ok"}


# ==========================================================================
# FORMATOS DE IMPRESIÓN (editor de plantillas)
# ==========================================================================
class FormatoInput(BaseModel):
    name: str = "Formato"
    page: dict = {"size": "A4", "orientation": "portrait"}
    elements: List[dict] = []


@api_router.get("/formatos")
async def listar_formatos():
    cur = db.formatos_impresion.find({}, {"_id": 0}).sort("created_at", 1)
    return [f async for f in cur]


@api_router.post("/formatos")
async def crear_formato(data: FormatoInput):
    doc = {
        "id": new_id(),
        "name": data.name,
        "page": data.page,
        "elements": data.elements,
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.formatos_impresion.insert_one(dict(doc))
    doc.pop("_id", None)
    return doc


@api_router.put("/formatos/{fid}")
async def actualizar_formato(fid: str, data: FormatoInput):
    res = await db.formatos_impresion.update_one(
        {"id": fid},
        {"$set": {"name": data.name, "page": data.page, "elements": data.elements, "updated_at": now_iso()}},
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Formato no encontrado")
    doc = await db.formatos_impresion.find_one({"id": fid}, {"_id": 0})
    return doc


@api_router.delete("/formatos/{fid}")
async def eliminar_formato(fid: str):
    await db.formatos_impresion.delete_one({"id": fid})
    return {"ok": True}




app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup_seed():
    # Índices
    try:
        await db.users.create_index("email", unique=True)
        await db.licencias.create_index("license_key", unique=True)
    except Exception as e:
        logger.warning(f"Index warning: {e}")
    # Admin idempotente
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@nexopro.com").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "Admin1234!")
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        await db.users.insert_one({
            "id": new_id(), "email": admin_email, "password_hash": hash_password(admin_password),
            "name": "Administrador", "role": "admin", "created_at": now_iso(),
        })
        logger.info("Admin creado")
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})
    # Licencia demo idempotente
    demo_key = os.environ.get("DEMO_LICENSE_KEY", "NEXO-DEMO-0001")
    if not await db.licencias.find_one({"license_key": demo_key}):
        await db.licencias.insert_one({
            "id": new_id(), "license_key": demo_key, "empresa": "Empresa Demo SL",
            "email": "demo@empresa.es", "telefono": "", "precio_mensual": 29,
            "estado": "activa", "ultimo_pago": None, "proximo_pago": None,
            "notas": "Licencia de demostración", "created_at": now_iso(),
        })


    # Usuario administrador del TALLER (login del ERP) idempotente
    try:
        await db.app_users.create_index("email", unique=True)
    except Exception as e:
        logger.warning(f"Index app_users warning: {e}")
    if await db.app_users.count_documents({}) == 0:
        app_admin_email = os.environ.get("APP_ADMIN_EMAIL", "administrador@taller.com").lower()
        app_admin_pwd = os.environ.get("APP_ADMIN_PASSWORD", "Taller1234!")
        await db.app_users.insert_one({
            "id": new_id(), "nombre": "Administrador", "email": app_admin_email,
            "password_hash": hash_password(app_admin_pwd), "role": "admin", "activo": True,
            "failed_attempts": 0, "locked_until": None, "totp_secret": None,
            "totp_enabled": False, "must_change_password": True, "last_login": None,
            "created_at": now_iso(),
        })
        logger.info("Usuario admin del taller creado")

    # Contador de referencias de artículo (evita colisión con los existentes)
    if not await db.counters.find_one({"_id": "articulo_ref"}):
        existentes = await db.articulos.count_documents({})
        await db.counters.insert_one({"_id": "articulo_ref", "seq": existentes})

    # Scheduler de recordatorios automáticos de citas (cada 30 min)
    try:
        if not _scheduler.running:
            _scheduler.add_job(_job_recordatorios, "interval", minutes=30, id="recordatorios",
                               replace_existing=True, next_run_time=datetime.now() + timedelta(seconds=60))
            _scheduler.start()
            logger.info("Scheduler de recordatorios iniciado")
    except Exception as e:
        logger.warning(f"No se pudo iniciar el scheduler: {e}")


@app.on_event("shutdown")
async def shutdown_db_client():
    try:
        if _scheduler.running:
            _scheduler.shutdown(wait=False)
    except Exception:
        pass
    client.close()
